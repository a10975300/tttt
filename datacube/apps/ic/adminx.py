from django.contrib import messages
from django.http import HttpResponseRedirect

import xadmin
from .models import ICs,IC_Report_Import_Records
from xadmin.filters import MultiSelectFieldListFilter
from xadmin.layout import Main, Fieldset, Side, Row
from import_export import resources


class ICsAdmin(object):

    list_display = [
        "ic_IC_Function",
        "ic_Vendor",
        "ic_Vendor_PN",
        "ic_Description",
        "ic_Part_Owner",
        "ic_HP_PN",
        "ic_IC_qty",
        #"ic_Full_Feature",
        #"ic_Defeature",
        "ic_Vendor_2nd",
        #"ic_Vendor_PN_2nd",
        #"ic_Vendor_3rd",
        #"ic_Vendor_PN_3rd",
        #"ic_Vendor_4th",
        #"ic_Vendor_PN_4th",
        #"ic_cratedate"
        ]
    list_filter = [
        ("ic_Cycle", MultiSelectFieldListFilter),
        ("ic_Project_name", MultiSelectFieldListFilter),
        ("ic_HW_PM", MultiSelectFieldListFilter),
        ("ic_ODM", MultiSelectFieldListFilter),
        ("ic_Segment", MultiSelectFieldListFilter),
        ("ic_Chipset_Type", MultiSelectFieldListFilter),
        ("ic_Phase", MultiSelectFieldListFilter),
        ("ic_IC_Function", MultiSelectFieldListFilter),
        ("ic_Vendor", MultiSelectFieldListFilter),
        ("ic_Vendor_PN", MultiSelectFieldListFilter),
        ("ic_Description", MultiSelectFieldListFilter),
        ("ic_Part_Owner", MultiSelectFieldListFilter),
        ("ic_HP_PN", MultiSelectFieldListFilter),
        ("ic_IC_qty", MultiSelectFieldListFilter),
        ("ic_Full_Feature", MultiSelectFieldListFilter),
        ("ic_Defeature", MultiSelectFieldListFilter),
        ("ic_Vendor_2nd", MultiSelectFieldListFilter),
        ("ic_Vendor_PN_2nd", MultiSelectFieldListFilter),
        #("ic_Vendor_3rd", MultiSelectFieldListFilter),
        #("ic_Vendor_PN_3rd", MultiSelectFieldListFilter),
        #("ic_Vendor_4th", MultiSelectFieldListFilter),
        #("ic_Vendor_PN_4th", MultiSelectFieldListFilter),
        #("ic_cratedate"),
        ]

    search_fields = [
        "ic_Cycle",
        "ic_Project_name",
        "ic_HW_PM",
        "ic_ODM",
        "ic_Segment",
        "ic_Chipset_Type",
        "ic_Phase",
        "ic_IC_Function",
        "ic_Vendor",
        "ic_Vendor_PN",
        "ic_Description",
        "ic_Part_Owner",
        "ic_HP_PN",
        "ic_IC_qty",
        "ic_Full_Feature",
        "ic_Defeature",
        "ic_Vendor_2nd",
        "ic_Vendor_PN_2nd",
        "ic_Vendor_3rd",
        "ic_Vendor_PN_3rd",
        "ic_Vendor_4th",
        "ic_Vendor_PN_4th",
        ]

    list_display_links = ["ic_IC_Function"]
    list_per_page = 50
    date_hierarchy = 'ic_cratedate'
    list_editable = 'ic_IC_Function'
    model_icon = 'fa fa-bandcamp'

    # configuration for import and export data
    class Issue_import_export_Resource(resources.ModelResource):
        class Meta:
            model = ICs
    import_export_args = {'export_resource_class':Issue_import_export_Resource}
    import_excel = True

    # rewrite post method for getting excel ic reprot
    def post(self, request, *args, **kwargs):
        currentuser = self.request.user  # get current user

        if 'excel' in request.FILES:
            ic_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(ic_report, data_only=True, keep_vba=False, keep_links=False)
            #取得IC EXCEL Sheet name
            sheet_names = workbook.sheetnames
            ic_sheet_name = sheet_names[0]

            ic_sheet = workbook[ic_sheet_name]
            project_name = ic_sheet.cell(2, 2).value
            cycle_name = ic_sheet.cell(1, 2).value
            # check if the product is the first upload or repeat to upload(don't support repeated upload)
            record = IC_Report_Import_Records.objects.filter(import_Project_name=project_name, import_Cycle=cycle_name)

            if record:  # repeated uploads

                # raise a message on webpage
                messages.error(request,
                               "Failed to import data! error message: [Project:{} -- {}'IC data already imported, repeated uploads do not supported]".format(
                                   record[0].import_Project_name, record[0].import_Cycle))
                """

                # send an alert by email
                contents = "Failed to import data!  error message: [Project:{} -- {}'IC data already imported, repeated uploads do not supported]".format(
                    record[0].import_Project_name, record[0].import_Cycle)
                Notification().send_by_email(
                    [str(currentuser)],
                    currentuser,
                    contents,
                    record[0].import_Project_namee,  # Project name
                    record[0].import_Cycle,  # Cycle
                    "IC report upload Failed",  # email title
                    None,
                    None, )
                """

            else:  # upload field
                from utils.excel_report_parser import IcReportParser
                result = IcReportParser().parse(request, workbook, currentuser)
                # raise a message on webpage
                if len(result) != 1:
                    messages.success(request, "{}-{}'s IC data was imported successfully.".format(result[0], result[1]))
                    """
                    #   send an alert by email
                    contents = "{}-{} IC data was imported successfully.".format(result[0],result[1])
                    Notification().send_by_email(
                                                [str(currentuser)],
                                                currentuser,
                                                contents,
                                                result[0],                                  # Project name
                                                result[1],                                  # cycle
                                                "IC report uploaded Successfully",  # email title
                                                None,                                  # 返回data
                                                None,)                                 # odm name
                    """
                else:
                    messages.error(request,"Failed, error: {}, Please use check tool to check the report first!".format(result[0]))
            workbook.close()
            return HttpResponseRedirect('/scpe/ic/ics/')
        return super(ICsAdmin, self).post(request, *args, **kwargs)


    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('ic_Cycle', 'ic_Project_name'),
                         Row('ic_HW_PM','ic_ODM'),
                         Row('ic_Segment', 'ic_Chipset_Type'),
                         Row('ic_Phase'),
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('ic_IC_Function','ic_IC_qty'),
                         Row('ic_Description'),
                         Row('ic_Defeature', 'ic_Full_Feature'),
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('ic_Vendor_2nd', 'ic_Vendor_PN_2nd'),
                         Row('ic_Vendor_3rd', 'ic_Vendor_PN_3rd'),
                         Row('ic_Vendor_4th', 'ic_Vendor_PN_4th'),
                         css_class='unsort no_title'
                         ),
                ),
            Side(
                Fieldset('',
                         'ic_Vendor', 'ic_Vendor_PN', 'ic_Part_Owner', 'ic_HP_PN',
                         css_class='unsort no_title'
                         ),
            ),
            )
        return super(ICsAdmin, self).get_form_layout()

xadmin.site.register(ICs, ICsAdmin)