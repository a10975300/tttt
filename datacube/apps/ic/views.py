
from django.contrib.auth.models import User
from .models import ICs
from openpyxl.cell import MergedCell

#ic excel report upload
class IC_Import_to_Database(object):
    """
    ic report upload to database,
    get excel contents from each cell then write to mysql database
    """
    def excel_to_database(self, workbook,ic_sheet_name):
        """
        从excel表ic sheet里面读取产品相关的信息如：产品类别，segment,产品名称，产品尺寸，ODM名称 以及build year
        """
        ic_sheet = workbook[ic_sheet_name]
        ic_Cycle = ic_sheet.cell(1, 2).value
        ic_Project_name = ic_sheet.cell(2, 2).value

        rows_number = ic_sheet.max_row  #get total row numbers
        empty_time = 0 #檢查最大行數之後，有三次空白，才為真正的最大行數

        for row in range (9, rows_number):
            checkingItem = ic_sheet.cell(row, 1).value
            checking_Vendor_PN = ic_sheet.cell(row, 3).value
            if checkingItem:
                NewItemHandler.ic_review_item_handle(self, ic_sheet,checkingItem,checking_Vendor_PN,row)
                empty_time = 0
            else:
                empty_time=empty_time+1
                if empty_time == 5:
                    break
        NewItemHandler.ic_import_record_handle(self,ic_Cycle,ic_Project_name)


# automatically handle product, new check-item related information
class NewItemHandler(object):

    # creare or update new item
    def ic_review_item_handle(self, ic_sheet,checkingItem,checking_Vendor_PN,row):

        #檢查第7欄位，有沒有合併欄位
        cell_merg = ic_sheet.cell(row=row, column=7)
        if isinstance(cell_merg, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in ic_sheet.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                if cell_merg.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell_merg = ic_sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        IC_qty_merg = cell_merg.value

        # 檢查第5欄位，替換B/S
        Part_Owner_clearn=ic_sheet.cell(row, 5).value
        if Part_Owner_clearn == 'B/S':
            Part_Owner_clearn = 'BS'


        new_item = ICs.objects.filter(ic_IC_Function=checkingItem,ic_Vendor_PN=checking_Vendor_PN)
        new_item.update_or_create(defaults={
            'ic_Cycle': ic_sheet.cell(1, 2).value,
            'ic_Project_name': ic_sheet.cell(2, 2).value,
            'ic_HW_PM': ic_sheet.cell(3, 2).value,
            'ic_ODM': ic_sheet.cell(4, 2).value,
            'ic_Segment': ic_sheet.cell(5, 2).value,
            'ic_Chipset_Type': ic_sheet.cell(6, 2).value,
            'ic_Phase': ic_sheet.cell(7, 2).value,
            'ic_IC_Function': ic_sheet.cell(row, 1).value,
            'ic_Vendor': ic_sheet.cell(row, 2).value,
            'ic_Vendor_PN': ic_sheet.cell(row, 3).value,
            'ic_Description': ic_sheet.cell(row, 4).value,
            'ic_Part_Owner': Part_Owner_clearn,
            'ic_HP_PN': ic_sheet.cell(row, 6).value,
            'ic_IC_qty': IC_qty_merg,
            'ic_Full_Feature': ic_sheet.cell(row, 8).value,
            'ic_Defeature': ic_sheet.cell(row, 9).value,
            'ic_Vendor_2nd': ic_sheet.cell(row, 10).value,
            'ic_Vendor_PN_2nd': ic_sheet.cell(row, 11).value,
            'ic_Vendor_3rd': ic_sheet.cell(row, 12).value,
            'ic_Vendor_PN_3rd': ic_sheet.cell(row, 13).value,
            'ic_Vendor_4th': ic_sheet.cell(row, 14).value,
            'ic_Vendor_PN_4th': ic_sheet.cell(row, 15).value,

        })

    # create or update an record when ic report import sucessfully
    def ic_import_record_handle(self, ic_Cycle,ic_Project_name):
        from .models import IC_Report_Import_Records

        new_item = IC_Report_Import_Records.objects.filter(import_Cycle=ic_Cycle,import_Project_name=ic_Project_name)
        new_item.update_or_create(defaults={
            'user_id': User.objects.filter(username=self.request.user)[0].id,
            'import_Cycle': ic_Cycle,
            'import_Project_name': ic_Project_name,
        })


