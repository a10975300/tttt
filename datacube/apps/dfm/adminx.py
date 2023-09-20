from xadmin.filters import MultiSelectFieldListFilter
from django.contrib import messages
from django.http import HttpResponseRedirect

import xadmin
from .models import Dfm_General_Checklist,Dfm_Review_Result,DFM_Report_Import_Records
from xadmin.layout import Main, Fieldset, Side, Row
from import_export import resources
from utils.notification import Notification

# define dfm user interface
class DfmAdmin(object):
    list_display = [
        "dfm_item_item_no",
        "dfm_item_desc",
        "dfm_item_priority",
        "dfm_assembly_level",
        "dfm_item_attributes",
        "dfm_item_version",
        "dfm_item_notes",
        "crate_date"
        ]
    list_filter = [
        ("dfm_assembly_level", MultiSelectFieldListFilter),
        # ("dfm_item_desc", MultiSelectFieldListFilter),
        ("dfm_item_priority", MultiSelectFieldListFilter),
        ("dfm_item_attributes", MultiSelectFieldListFilter),
        ("dfm_item_version", MultiSelectFieldListFilter),
        ("dfm_item_notes", MultiSelectFieldListFilter),
        ("crate_date"),
        ]
    search_fields = ["dfm_item_desc","dfm_assembly_level","dfm_item_priority","dfm_assembly_level","dfm_item_attributes","dfm_item_version","dfm_item_notes"]
    list_display_links = ["dfm_item_desc"]
    list_per_page = 50
    date_hierarchy = 'crate_date'
    list_editable = ('dfm_item_desc', 'dfm_item_attributes')
    ordering = ['dfm_item_item_no']
    model_icon = 'fa fa-list-alt'

    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('dfm_item_item_no', 'dfm_item_priority'),
                         'dfm_item_desc',
                         Row("dfm_assembly_level", 'dfm_item_attributes'),
                         'crate_date',
                         css_class='unsort no_title'
                         ),
            ),
            Side(
                Fieldset(('Additional info'),
                         'dfm_item_version', 'dfm_item_notes',
                         ),
            )
        )
        return super(DfmAdmin, self).get_form_layout()

class Dfm_Review_Admin(object):
    list_display = [
                    "dfm_review_item_desc",
                    "dfm_product",
                    "dfm_product_nud",
                    "dfm_product_cnc",
                    "dfm_product_si",
                    "dfm_product_pv",
                    "dfm_product_mv",
                    # "dfm_product_part_category",
                    "dfm_product_issue_symptom",
                    "dfm_product_design_structures",
                    # "dfm_product_nonmacth_item",
                    # "dfm_product_odm_actions",
                    "dfm_product_solution_category",
                    ]
    list_filter = [
        ("dfm_product__ProductName", MultiSelectFieldListFilter),
        ("dfm_product_part_category", MultiSelectFieldListFilter),
        ("dfm_product_issue_symptom", MultiSelectFieldListFilter),
        ("dfm_product_design_structures", MultiSelectFieldListFilter),
        ("dfm_product_solution_category", MultiSelectFieldListFilter),
        ("dfm_product_nud", MultiSelectFieldListFilter),
        ("dfm_product_cnc", MultiSelectFieldListFilter),
        ("dfm_product_si", MultiSelectFieldListFilter),
        ("dfm_product_pv", MultiSelectFieldListFilter),
        ("dfm_product_mv", MultiSelectFieldListFilter),
        ]
    search_fields = ["dfm_review_item_desc__dfm_item_desc","dfm_product__ProductName","dfm_product_issue_symptom",
                     "dfm_product_design_structures","dfm_product_solution_category"]
    list_display_links = ["dfm_review_item_desc"]
    list_per_page = 10
    date_hierarchy = 'crate_date'
    list_editable = ('')
    model_icon = 'fa fa-check-square-o'

    ### Dfm_review_product 注册导入导出功能
    class Issue_import_export_Resource(resources.ModelResource):
        class Meta:
            model = Dfm_Review_Result
    import_export_args = {'export_resource_class':Issue_import_export_Resource}
    import_excel = True

    # rewrite post method for getting excel DFM reprot
    def post(self, request, *args, **kwargs):
        currentuser = self.request.user # get current user

        if 'excel' in request.FILES:
            dfm_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(dfm_report, data_only=True, keep_vba=False, keep_links=False)

            dfm_sheet = workbook['DFm ']
            product_name = dfm_sheet .cell(6, 13).value
            report_version = dfm_sheet.cell(2, 2).value  # get dfm report template version

            if report_version and report_version.find("Ver:1.63") != -1: #report_version == Ver:1.63
                dfm_stage = "MV"
                report_version = "Ver:1.63"
                record = None

            elif report_version and report_version.find("Ver:2.0") != -1:  # report_version == Ver:2.0
                # check if the product and dfm stage is the first upload or repeat to upload(don't support repeated upload)
                cnc = dfm_sheet.cell(32, 4).value
                si = dfm_sheet.cell(32, 5).value
                pv = dfm_sheet.cell(32, 6).value
                mv = dfm_sheet.cell(32, 7).value
                report_version = "Ver:2.0"

                if mv:  # mv
                    dfm_stage = "MV"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y",
                                                                      import_stage_pv="Y", import_stage_mv="Y")
                elif pv:  # pv
                    dfm_stage = "PV"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y",
                                                                      import_stage_pv="Y")
                elif si:  # si
                    dfm_stage = "SI"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y")
                else:  # cnc
                    dfm_stage = "CNC"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y")

            else:  # report_version == Ver:3.0a
                # check if the product and dfm stage is the first upload or repeat to upload(don't support repeated upload)
                cnc = dfm_sheet.cell(32, 5).value
                si = dfm_sheet.cell(32, 6).value
                pv = dfm_sheet.cell(32, 7).value
                mv = dfm_sheet.cell(32, 8).value
                report_version = "Ver:3.0a"

                if mv:  # mv
                    dfm_stage = "MV"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y",
                                                                      import_stage_pv="Y", import_stage_mv="Y")
                elif pv:  # pv
                    dfm_stage = "PV"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y",
                                                                      import_stage_pv="Y")
                elif si:  # si
                    dfm_stage = "SI"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y", import_stage_si="Y")
                else:  # cnc
                    dfm_stage = "CNC"
                    record = DFM_Report_Import_Records.objects.filter(import_product_name=product_name,
                                                                      import_stage_cnc="Y")


            if record :#repeated uploads

                # raise a message on webpage
                messages.error(request,
                               "Failed to import data! error message: [ProductId:{} -- {} -- {}'dfm data already imported, repeated uploads do not supported]".format(
                                   record[0].id, record[0].import_product_name,dfm_stage))
                """
                # send an alert by email
                contents = "Failed to import data!  error message: [ProductId:{} -- {} -- {}'dfm data already imported, repeated uploads do not supported]".format(
                    record[0].id, record[0].import_product_name,dfm_stage)
                Notification().send_by_email(
                    [str(currentuser)],
                    currentuser,
                    contents,
                    record[0].import_product_name,  # platform name
                    dfm_stage,  # dfm_stage
                    "dfm report upload Failed",  # email title
                    None,
                    None, )
                """
            elif report_version == "Ver:1.63": #just for 1.63 upload
                from utils.excel_report_parser import DfmReportParser
                result = DfmReportParser().parse(request, workbook, currentuser,report_version)
                # raise a message on webpage
                if len(result) != 1:
                    messages.success(request, "{}-{}'s dfm report was imported successfully.".format(result[0],result[1]))
                else:
                    messages.error(request,"Failed, error: {}, Please use check tool to check the report first!".format(result[0]))

            else: #upload field
                from utils.excel_report_parser import DfmReportParser
                result = DfmReportParser().parse(request, workbook, currentuser,report_version)
                # raise a message on webpage
                if len(result) != 1:
                    messages.success(request, "{}-{}'s dfm report was imported successfully.".format(result[0],result[1]))

                    #   send an alert by email
                    contents = "{}-{} dfm report was imported successfully.".format(result[0],result[1])
                    Notification().dfm_send_by_email(
                                                user =[str(currentuser)],
                                                platform_name =result[0],
                                                dfm_stage=result[1],
                                                odm_name=result[2],
                                                subject=contents
                                                )
                else:
                    messages.error(request,"Failed, error: {}, Please use check tool to check the report first!".format(result[0]))
            workbook.close()
            return HttpResponseRedirect('/scpe/dfm/dfm_review_result/')
        return super(Dfm_Review_Admin, self).post(request, *args, **kwargs)

    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('dfm_review_item_no', 'dfm_product'),
                         Row('dfm_review_item_desc'),
                         Row('dfm_product_nud', 'dfm_product_part_category'),
                         Row('dfm_product_issue_symptom', 'dfm_product_design_structures'),

                         css_class='unsort no_title'
                    ),
                Fieldset('',
                         Row('dfm_product_nonmacth_item', 'dfm_product_odm_actions'),
                         'dfm_product_solution_category',
                         css_class='unsort no_title'
                    ),
            ),
            Side(
                Fieldset('Teardown Review Result',
                         Row('dfm_product_cnc', 'dfm_product_si'),
                         Row('dfm_product_pv','dfm_product_mv'),
                         'photo'
                         ),
            )
        )
        return super(Dfm_Review_Admin, self).get_form_layout()

class Dfm_Upload_Records_Admin(object):
    list_display = ["user","import_product_name","import_stage_cnc","import_stage_si","import_stage_pv","import_stage_mv","import_date"]
    list_display_links =["import_product_name"]

    def save_models(self):
        self.new_obj.user = self.request.user
        return super().save_models()

#后台管理项目注册
xadmin.site.register(Dfm_General_Checklist, DfmAdmin)
xadmin.site.register(Dfm_Review_Result, Dfm_Review_Admin)
xadmin.site.register(DFM_Report_Import_Records, Dfm_Upload_Records_Admin)