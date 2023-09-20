import datetime
import requests
from PIL import Image
from io import BytesIO

from django.contrib.auth.models import User
from npi.models import Issue, SymptomCategory_First, SymptomCategory_Second,DesktopIssue
from dfm.models import Dfm_General_Checklist, Dfm_Review_Result,DFM_Report_Import_Records
from product.models import Products
from dfa.models import Dfa_Review_Result, Dfa_Report_Import_Records
from ic.models import ICs,IC_Report_Import_Records


from openpyxl.cell import MergedCell
from npi.forms import IssueModelForm,DesktopIssueModelForm
from dfa.forms import DfaModelForm
from ic.forms import ICModelForm
from dfm.forms import DfmIssueModelForm

class SafelaunchParser():
    """
    parse safelaunch report
    """
    def parse(self, request, workbook, currentuser):
        """ check safe launch template version then assign the right method()"""
        report_version = workbook['0-Summary Data'].cell(1,1).value   # get safe launch report template version from summary sheet
        if report_version == 'V2.2f (2022/12/09)' or report_version == "V3a (2023/03/02)":
            result = self.new_template(request=request, workbook=workbook, currentuser=currentuser)
            return result

        else: # report_version == 'V2.1 (2022/10/25)' or other old template
            result = self.old_template(request=request, workbook=workbook, currentuser=currentuser)
            return result

    # parse new template V2.2e (2022/12/06)
    def new_template(self,request, workbook, currentuser):
        global issue_list

        summary_sheet = workbook['0-Summary Data']
        odm_name = summary_sheet.cell(3, 2).value             # get odm name from summary sheet
        segment = summary_sheet.cell(4, 2).value              # get product segment from summary sheet
        product_name = summary_sheet.cell(5, 2).value         # get product name from summary sheet
        build_stage = summary_sheet.cell(7, 2).value          # get build stage from summary sheet

        sheet_names = workbook.sheetnames
        data = {}  # initial a null dict for collecting issue qty by each sheet
        issue_list = [] # initial a null list to store issues temporarily

        for sheet_name in sheet_names[2:11]:
            # count issue qty by each sheet (default is 0)
            issue_qty_by_sheet = 0
            issue_qty_gating = 0
            issue_qty_tracking = 0
            issue_qty_close = 0
            issue_qty_fix_in_next_stage = 0

            sheet = workbook[sheet_name]  # 通过sheet名称指定sheet
            rows_number = sheet.max_row  # get total row numbers
            # 调用get_picture_from_excels方法 处理sheet里面的图片，返回每张图片的具体行数和列数
            if sheet._images:
                self.result, self.buildstage, self.platform = self.get_pictures_from_excel(sheet, product_name, build_stage, sheet_name)
            else:
                self.result = {}
            # write excel contents into database by for row in range(4, rows_number):
            for row in range(5, rows_number, 4):
                if sheet.cell(row, 3).value :  # check if have contents in excel row? if yes, execute below codings; if no, break
                    unit_size = sheet.cell(row, 2).value
                    if not unit_size:
                        error = {
                            "sheet_name": sheet_name,
                            "row": row,
                            "message": "The column of unit_size is blank"  # list(result.keys())[0]
                        }
                        return [error]
                    productIds = self.create_new_product(unit_size, build_stage, odm_name, segment, currentuser)  # 调用函数 new_product 判断数据库是否已有该产品，
                    for productId in productIds:
                        issueCateory_1 = sheet.cell(row, 4).value  # write issue interaction
                        if not issueCateory_1:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message": "Issue interaction is blank"  # list(result.keys())[0]
                            }
                            return [error]
                        category_1_id = self.new_category_1(issueCateory_1)

                        issueCategory_2 = sheet.cell(row, 5).value  # write issue symptom
                        if not issueCategory_2:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message": "Issue symptom is blank"  # list(result.keys())[0]
                            }
                            return [error]
                        category_2_id = self.new_category_2(issueCategory_2, category_1_id, issueCateory_1)

                        pre_build_qty = (0 if (sheet.cell(row, 7).value) == None else (sheet.cell(row, 7).value))
                        pre_build_defcet_qty = (0 if (sheet.cell(row, 8).value) == None else (sheet.cell(row, 8).value))
                        mini1_build_qty = (0 if (sheet.cell(row + 1, 7).value) == None else (sheet.cell(row + 1, 7).value))
                        mini1_build_defcet_qty = (0 if (sheet.cell(row + 1, 8).value) == None else (sheet.cell(row + 1, 8).value))
                        mini2_build_qty = (0 if (sheet.cell(row + 2, 7).value) == None else (sheet.cell(row + 2, 7).value))
                        mini2_build_defcet_qty = (0 if (sheet.cell(row + 2, 8).value) == None else (sheet.cell(row + 2, 8).value))
                        balance_qty = (0 if (sheet.cell(row + 3, 7).value) == None else (sheet.cell(row + 3, 7).value))
                        balance_defcet_qty = (0 if (sheet.cell(row + 3, 8).value) == None else (sheet.cell(row + 3, 8).value))

                        # get and write pic path
                        if (sheet_name + str(11) + str(row)) in self.result:
                            photo = "issue/images/{}/{}/{}".format(self.platform, self.buildstage, self.result.get(sheet_name + str(11) + str(row)))
                        else:
                            photo = ""
                        # get obs
                        obs_ = sheet.cell(row, 21).value
                        if obs_ == '' or obs_ == 'N/A' or obs_ == 'N' or obs_ == 0:
                            obs = 'NA'
                        else:
                            obs = obs_
                        # get error code
                        error = sheet.cell(row, 23).value  # write issue error code number
                        if error == '' or error == "N/A" or error == 'N' or error == 0:
                            errorCode = 'NA'
                        else:
                            errorCode = error

                        post_data = {
                            "platformName_id":productId,
                            "processName":sheet_name,
                            "issue_desc":sheet.cell(row, 3).value,
                            "issue_interaction_id":category_1_id,
                            "issue_symptom_id":category_2_id,
                            "impact_scope":sheet.cell(row, 6).value,
                            "pre_build_qty":pre_build_qty,
                            "pre_build_defcet_qty": pre_build_defcet_qty,
                            "mini_build_qty": mini1_build_qty,
                            "mini_build_defcet_qty": mini1_build_defcet_qty,
                            "mini2_build_qty": mini2_build_qty,
                            "mini2_build_defcet_qty": mini2_build_defcet_qty,
                            "balance_qty": balance_qty,
                            "balance_defcet_qty": balance_defcet_qty,
                            "input_qty": pre_build_qty + mini1_build_qty + mini2_build_qty + balance_qty,
                            "defect_qty": pre_build_defcet_qty + mini1_build_defcet_qty + mini2_build_defcet_qty + balance_defcet_qty,
                            "fail_rate_stage": "NA",
                            "sn":sheet.cell(row, 10).value,
                            "sku":sheet.cell(row, 11).value,
                            "photo":photo,

                            "issue_analysis": sheet.cell(row, 13).value.replace("\n", "<br/>"),
                            "root_cause": sheet.cell(row, 14).value.replace("\n", "<br/>"),
                            "root_cause_category": sheet.cell(row, 15).value,
                            "short_term": sheet.cell(row, 16).value.replace("\n", "<br/>"),
                            "short_term_category": sheet.cell(row, 17).value,
                            "long_term": sheet.cell(row, 18).value.replace("\n", "<br/>"),
                            "long_term_category": sheet.cell(row, 19).value,
                            "status": sheet.cell(row, 20).value,
                            "obs": obs,
                            "duplicate": sheet.cell(row, 22).value ,
                            "errorCode": errorCode,
                            "repeating": sheet.cell(row, 24).value,
                            "repeatingstage": sheet.cell(row, 25).value,
                            "buildstage": sheet.cell(row, 26).value,
                            "owner": sheet.cell(row, 27).value,
                            "updatedate": sheet.cell(row, 28).value,
                            "obs_link": "http://pulsarweb.twn.hp.com/Nebula//ReportRunner/RunDetailWithObsIds?observationIds={}".format(obs),
                        }

                        #  call issue_validate to verify issue contents
                        result = self.issue_validate(post_data, productId, category_1_id, category_2_id)
                        if result == "pass":
                            # count issue quantity
                            issue_qty_by_sheet += 1
                            if sheet.cell(row, 20).value == "Gating":
                                issue_qty_gating += 1
                            elif sheet.cell(row, 20).value == "Tracking":
                                issue_qty_tracking += 1
                            elif sheet.cell(row, 20).value == "Close":
                                issue_qty_close += 1
                            elif sheet.cell(row, 20).value == "Fix in next phase":
                                issue_qty_fix_in_next_stage += 1
                            else:
                                issue_qty_tracking += 1
                        else:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message":result  #list(result.keys())[0]
                            }
                            return [error]

            data[sheet_name] = {
                "toatl_issue_qty": issue_qty_by_sheet,
                "gating": issue_qty_gating,
                "tracking": issue_qty_tracking,
                "close": issue_qty_close,
                "fix_in_next_stage": issue_qty_fix_in_next_stage
            }

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            # import logs
            self.safelaunch_import_record(product_name, build_stage, currentuser)
            # log entries
            self.log_entries(request, product_name, build_stage, currentuser)
            return [product_name, build_stage, odm_name, data]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # parse old template V2.1 (2022/10/25)
    def old_template(self, request, workbook, currentuser):
        summary_sheet = workbook['0-Summary Data']
        odm_name = summary_sheet.cell(3, 2).value  # get odm name from summary sheet
        segment = summary_sheet.cell(4, 2).value  # get product segment from summary sheet
        product_name = summary_sheet.cell(5, 2).value  # get product name from summary sheet
        build_stage = summary_sheet.cell(7, 2).value  # get build stage from summary sheet

        sheet_names = workbook.sheetnames
        data = {}  # initial a null dict for collecting issue qty by each sheet
        for sheet_name in sheet_names[2:11]:
            # count issue qty by each sheet (default is 0)
            issue_qty_by_sheet = 0
            issue_qty_gating = 0
            issue_qty_tracking = 0
            issue_qty_close = 0
            issue_qty_fix_in_next_stage = 0

            sheet = workbook[sheet_name]  # 通过sheet名称指定sheet
            rows_number = sheet.max_row  # get total row numbers
            # 调用get_picture_from_excels方法 处理sheet里面的图片，返回每张图片的具体行数和列数
            if sheet._images:
                self.result, self.buildstage, self.platform = self.get_pictures_from_excel(sheet, product_name,
                                                                                           build_stage, sheet_name)
            else:
                self.result = {}
            # write excel contents into database by for row in range(4, rows_number):
            for row in range(5, rows_number):
                unit_size = sheet.cell(row, 2).value

                if unit_size and sheet.cell(row, 3).value:  # check if have contents in excel row? if yes, execute below codings; if no, break
                    productIds = self.create_new_product(unit_size, build_stage, odm_name, segment,
                                                         currentuser)  # 调用函数 new_product 判断数据库是否已有该产品，
                    for productId in productIds:
                        new_issue = Issue()
                        new_issue.platformName_id = productId
                        new_issue.processName = sheet_name  # write process name
                        new_issue.issue_desc = sheet.cell(row, 3).value  # write issue description
                        issueCategory_1 = sheet.cell(row, 4).value  # write issue interaction
                        category_1_id = self.new_category_1(issueCategory_1)
                        new_issue.issue_interaction_id = category_1_id
                        issueCategory_2 = sheet.cell(row, 5).value  # write issue symptom
                        category_2_id = self.new_category_2(issueCategory_2, category_1_id, issueCategory_1)
                        new_issue.issue_symptom_id = category_2_id
                        new_issue.impact_scope = sheet.cell(row, 6).value  # write issue impact scope

                        new_issue.pre_build_qty = 0
                        new_issue.pre_build_defcet_qty = 0
                        new_issue.mini_build_qty = 0          # mini-1 build qty
                        new_issue.mini_build_defcet_qty = 0   # mini-1 defect qty
                        new_issue.mini2_build_qty = 0         # mini-2 build qty
                        new_issue.mini2_build_defcet_qty = 0  # mini-2 defect qty
                        new_issue.balance_qty = 0             # balance build qty
                        new_issue.balance_defcet_qty = 0      # balance defect qty

                        # write total input quantity
                        new_issue.input_qty = (sheet.cell(row, 7).value)
                        # write defect quantity
                        new_issue.defect_qty = (sheet.cell(row, 8).value)

                        new_issue.fail_rate_stage = (sheet.cell(row, 9).value)  # write failure rate by every build stage
                        new_issue.sn = sheet.cell(row, 10).value  # write fail units S/N information
                        new_issue.sku = sheet.cell(row, 11).value  # write fail units SKU information

                        # get and write pic path
                        if (sheet_name + str(11) + str(row)) in self.result:
                            new_issue.photo = "issue/images/{}/{}/{}".format(self.platform, self.buildstage,
                                                                             self.result.get(
                                                                                 sheet_name + str(11) + str(row)))
                        else:
                            new_issue.photo = ""

                        new_issue.issue_analysis = sheet.cell(row, 13).value.replace("\n","<br/>")  # write issue analysis contents
                        new_issue.root_cause = sheet.cell(row, 14).value.replace("\n","<br/>")  # write root cause
                        new_issue.root_cause_category = sheet.cell(row, 15).value  # write root cause category
                        new_issue.short_term = sheet.cell(row, 16).value.replace("\n","<br/>")  # write short term solutions
                        new_issue.short_term_category = sheet.cell(row, 17).value  # write short term solution category
                        new_issue.long_term = sheet.cell(row, 18).value.replace("\n","<br/>")  # write long-term solutions
                        new_issue.long_term_category = sheet.cell(row, 19).value  # write long-term solution category
                        new_issue.status = sheet.cell(row, 20).value  # write issue status

                        obs = sheet.cell(row, 21).value
                        if obs == '' or obs == 'N/A' or obs == 'N' or obs == 0:
                            obs = 'NA'
                        new_issue.obs = obs  # write obs number

                        new_issue.duplicate = sheet.cell(row, 22).value  # write duplicate info
                        errorCode = sheet.cell(row, 23).value  # write issue error code number
                        if errorCode == '' or errorCode == "N/A" or errorCode == 'N' or errorCode == 0:
                            new_issue.errorCode = 'NA'
                        new_issue.errorCode = errorCode

                        new_issue.repeating = sheet.cell(row, 24).value  # write issue repeating info
                        new_issue.repeatingstage = sheet.cell(row, 25).value  # write issue repeated stages
                        new_issue.buildstage = sheet.cell(row, 26).value  # write issue build stage
                        new_issue.owner = sheet.cell(row, 27).value  # write issue owner
                        new_issue.updatedate = sheet.cell(row, 28).value
                        new_issue.obs_link = "http://pulsarweb.twn.hp.com/Nebula//ReportRunner/RunDetailWithObsIds?observationIds={}".format(
                            obs)  # write obs links

                        new_issue.save()

                        issue_qty_by_sheet += 1  # count issue qty by sheet
                        if sheet.cell(row, 20).value == "Gating":
                            issue_qty_gating += 1
                        elif sheet.cell(row, 20).value == "Tracking":
                            issue_qty_tracking += 1
                        elif sheet.cell(row, 20).value == "Close":
                            issue_qty_close += 1
                        elif sheet.cell(row, 20).value == "Fix in next phase":
                            issue_qty_fix_in_next_stage += 1
                        else:
                            issue_qty_tracking += 1

            data[sheet_name] = {
                "toatl_issue_qty": issue_qty_by_sheet,
                "gating": issue_qty_gating,
                "tracking": issue_qty_tracking,
                "close": issue_qty_close,
                "fix_in_next_stage": issue_qty_fix_in_next_stage
            }

        # import logs
        self.safelaunch_import_record(product_name, build_stage, currentuser)
        # log entries
        self.log_entries(request, product_name, build_stage, currentuser)
        return [product_name, build_stage, odm_name, data]

    # issue validate
    def issue_validate(self, post_data, productId, category_1_id, category_2_id):
        form = IssueModelForm(data=post_data)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.platformName_id = productId
            instance.issue_interaction_id = category_1_id
            instance.issue_symptom_id = category_2_id

            issue_list.append(post_data)
            return "pass"
        else:
            return form.errors.as_data()

    # save issues to database
    def issue_save(self, issue_list):
        for issue in issue_list:
            form = IssueModelForm(data=issue)
            instance = form.save(commit=False)
            instance.platformName_id = issue['platformName_id']
            instance.issue_interaction_id = issue['issue_interaction_id']
            instance.issue_symptom_id = issue['issue_symptom_id']
            instance.save()
        return "succeed"

    # handel date formate because that the excel 读出数字日期（44557）转换函数
    def date_conversion(self, date):
        delta = datetime.timedelta(days=date)
        newdate = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
        return datetime.datetime.strftime(newdate, '%Y-%m-%d')

    # automatically handle product infomation
    def create_new_product(self, unit_size, build_stage, odm_name, segment, currentuser):
        """
        check if this product already in the database 使用productName查询数据库里是否已有该产品,
            if yes, then return product id
            if no, create a new one and return the product id
        """
        product_id =[] # initial a null list to record product IDs

        user = User.objects.filter(username=currentuser)
        unit_size = unit_size.split(",") # convert unit_size to list by ","
        for product in unit_size:
            obj = Products.objects.filter(ProductName=product.strip())
            if obj:
                product_id.append(obj[0].id)
                Products.objects.filter(ProductName=product).update(ProductPhase=build_stage) # update build stage only
            else:
                new_product = Products.objects.create(
                    ProductName = product,
                    Product_Segment = segment,
                    ProductPhase = build_stage,
                    PartnerName = odm_name,
                    user = currentuser,
                    PE = [(res.first_name + " " + res.last_name) for res in user][0]
                )
                new_product.save()
                product_id.append(new_product.id)
        return product_id

    # automatically handle issue's category
    def new_category_1(self, issueCategory_1):
        """
        check if this issue category already in the database
            if yes, return id
            if no, create a new category_1 and also category_2
        """
        category_1 =  SymptomCategory_First.objects.filter(category_Name=issueCategory_1)
        if category_1:
            return category_1[0].id
        else:
            newCategory = SymptomCategory_First.objects.create(
                category_Name = issueCategory_1, # write category name
                category_Desc = "This category name is {} as a 1st issue category.".format(issueCategory_1)
            )
            newCategory.save()
            return newCategory.id

    def new_category_2(self, issueCategory_2, category_1_id, issueCategory_1):
        category_2 = SymptomCategory_Second.objects.filter(category_Name=issueCategory_2)
        if category_2:
            return category_2[0].id
        else:
            newCategory = SymptomCategory_Second.objects.create(
                category_FirstType_id = category_1_id,
                category_Name=issueCategory_2,  # write category name
                category_Desc="This is a 2nd issue category belongs to {}".format(issueCategory_1)
            )
            newCategory.save()
            return newCategory.id

    # get pictures from excel in each sheet
    def get_pictures_from_excel(self, sheet, product_name, build_stage, sheet_name):
        # remove symbols from platform name
        product_name = product_name.replace(" ", "").replace("/", "_").replace("'", "").replace("\"", "").strip()
        # initial a null dict for recording pictures location(stage, sheet, row, column...)
        picture_location_dict = {}
        stages = {
            "SI-1": "SI_1",
            "SI-2": "SI_2",
            "SI-3": "SI_3",
            "PV-1": "PV_1",
            "PV-2": "PV_2",
            "PV-3": "PV_3",
            "PRB/TLD/PVR": "PRB_TLD_PVR",
            "MV-1": "MV_1",
            "MV-2": "MV_2"
        }
        for image in sheet._images:
            img = Image.open(BytesIO(image._data()))
            img.convert("RGB")  # fix error of "OSError: cannot write mode RGBA as JPEG" ,
            # print(image._data())
            """
            png格式图片颜色为四通道RGBA，而jpg格式是三通道RGB, （RGBA是代表Red（红色）Green（绿色）Blue（蓝色）和Alpha的色彩空间。
            虽然它有的时候被描述为一个颜色空间，但是它其实仅仅是RGB模型的附加了额外的信息。alpha通道一般用作不透明度参数。）
            """
            # img.show()
            img_name = ("{}_" + stages.get(build_stage) + "_Sheet" + sheet_name[0] + "_col{}_row{}.jpg").format(
                product_name, image.anchor._from.col, image.anchor._from.row)
            # return
            if (image.anchor._from.col, image.anchor._from.row):
                picture_location_dict["{}{}{}".format(sheet_name, str(image.anchor._from.col), str(image.anchor._from.row + 1))] = img_name

                # 单线程 -- 调用pictureUpload方法 requests.post 从本地跨域提交图片到服务器
                self.pictureUpload(product_name, stages.get(build_stage), img_name, image._data())

                # 多线程 -- 调用pictureUpload方法 requests.post 从本地跨域提交图片到服务器
                # t = threading.Thread(target=self.pictureUpload, args=(product_name, stages.get(build_stage), img_name, image._data()))
                # t.start()

        return picture_location_dict, stages.get(build_stage), product_name

    # 通过requests.post模拟游览器提交图片和信息
    def pictureUpload(self, product_name, stage_name, img_name, image):
        #url = 'http://127.0.0.1:8000/pictures_upload/'  # this url for localhost test only
        url = 'http://15.36.145.93/pictures_upload/'    # this url for server
        myData = {
            "product_name": product_name,
            "stage_name": stage_name,
            "pic_name": img_name
        }
        myFiles = {
            "img": image
        }
        res = requests.post(url, files=myFiles, data=myData)
        res.close()

    # import logs
    def safelaunch_import_record(self, prodct_name, build_stage, currentuser):
        from npi.models import DataImportRecords
        # 记录数据导入信息
        importRecord = DataImportRecords.objects.create(
            import_product_name=prodct_name,
            import_product_phase=build_stage,
            user=currentuser
        )
        importRecord.save()

        # write a log into Log entries after safe-launch report uploaded

    def log_entries(self, request, product_name, build_stage, currentuser):
        from xadmin.models import Log
        Log.objects.create(user=currentuser,
                           ip_addr=request.META.get('REMOTE_ADDR'),
                           object_repr=str(object),
                           action_flag='',
                           message="{}-{} safelaunch report uploaded sucessfully".format(product_name, build_stage))

class DfaReportParser():
    """
    parse Dfa tear down report
    """
    def parse(self, request, workbook, currentuser):
        """ call dfa method()"""
        result = self.dfa_template(request=request, workbook=workbook, currentuser=currentuser)
        return result

    # parse dfa template
    def dfa_template(self,request, workbook, currentuser):
        dfa_sheet = workbook['Dashboard Page']
        product_name = dfa_sheet.cell(2, 13).value
        odm_name = dfa_sheet.cell(2, 7).value
        global issue_list
        issue_list = []  # initial a null list to store issues temporarily

        if not product_name:
            error = {
                "The column of product_name is blank"
            }
            return [error]
        product_id = self.new_product(product_name, odm_name, currentuser)

        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names[2:11]:#依序讀取excel的分頁
            sheet = workbook[sheet_name]  # 通过sheet名称指定sheet
            rows_number = sheet.max_row  # get total row numbers
            empty_time = 0  # 檢查最大行數之後，有三次空白，才為真正的最大行數
            for row in range(17, rows_number):
                checkingItem = sheet.cell(row, 3).value #object欄位 當作檢查
                if checkingItem:
                    dfa_object = (sheet.cell(row, 3).value).replace("\n", "\r\n")  # 使用.replace("\n", "\r\n")处理excel单元格里有换行问题 "\r\n"是textarea换行符
                    sheet_name=sheet_name.strip()#去除sheet name 的空白
                    post_data=self.dfa_item_handle(dfa_object,product_id,sheet,row,sheet_name)
                    empty_time = 0

                    #  call issue_validate to verify issue contents
                    result = self.issue_validate(post_data, product_id)
                    if result == "pass":
                        pass
                    else:
                        error = {
                            "sheet_name": sheet_name,
                            "row": row,
                            "message": result  # list(result.keys())[0]
                        }
                        return [error]

                else:
                    empty_time = empty_time+1
                    if empty_time == 3:
                        break

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            # call method of dfa_import_record_handle to create or update an record when the above steps finished
            self.dfa_import_record_handle(product_name, odm_name, currentuser)
            # log entries
            self.log_entries(request, product_name, odm_name, currentuser)
            return [product_name, odm_name]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # create or update product
    def new_product(self,product_name,odm_name,currentuser):
        """
        check if this product already in the database 使用productName查询数据库里是否已有该产品,
            if yes, then return product id
            if no, create a new one and return the product id
        """
        item = Products.objects.filter(ProductName=product_name,PartnerName=odm_name).update_or_create(defaults={
            'ProductName': product_name,
            'PartnerName': odm_name,
            'user': currentuser,
        })
        return item[0].id

    # handle dfa item
    def dfa_item_handle(self,dfa_object,product_id,sheet,row,sheet_name):

        #處理non_match_reason 和 odm_corrective_action 欄位
        non_match_reason = sheet.cell(row, 14).value
        odm_corrective_action = sheet.cell(row, 15).value
        non_match_reason_cnc = ""
        non_match_reason_si = ""
        non_match_reason_si2 = ""
        non_match_reason_pv = ""
        odm_corrective_action_cnc = ""
        odm_corrective_action_si = ""
        odm_corrective_action_si2 = ""
        odm_corrective_action_pv = ""

        if non_match_reason:
            non_match_reason = non_match_reason.replace("(CNC):", "#reason").replace("(SI):", "#reason").replace("(SI-2):", "#reason").replace("(PV):", "#reason")
            non_match_reason = non_match_reason.split('#reason')
            non_match_reason_cnc = non_match_reason[1]
            non_match_reason_si = non_match_reason[2]
            non_match_reason_si2 = non_match_reason[3]
            non_match_reason_pv = non_match_reason[4]

        if odm_corrective_action:
            odm_corrective_action = odm_corrective_action.replace("(CNC):", "#action").replace("(SI):","#action").replace("(SI-2):", "#action").replace("(PV):", "#action")
            odm_corrective_action = odm_corrective_action.split('#action')
            odm_corrective_action_cnc = odm_corrective_action[1]
            odm_corrective_action_si = odm_corrective_action[2]
            odm_corrective_action_si2 = odm_corrective_action[3]
            odm_corrective_action_pv = odm_corrective_action[4]

        post_data = {
            'dfa_product_id': product_id,
            'dfa_category': sheet_name,
            'dfa_object': dfa_object,
            'dfa_estimated_hc': sheet.cell(row, 4).value,
            'dfa_production_line': sheet.cell(row, 5).value,
            'dfa_cnc': sheet.cell(row, 6).value,
            'dfa_si_pv': sheet.cell(row, 7).value,
            'dfa_mv_mp': sheet.cell(row, 8).value,
            'dfa_HP_commodities_design': sheet.cell(row, 9).value,
            'dfa_odm_product_designs': sheet.cell(row, 10).value,
            'dfa_odm_packing_designs_kitting_box_design': sheet.cell(row, 11).value,
            'dfa_HP_packing_designs': sheet.cell(row, 12).value,
            'dfa_equipment_facilites_UPH_concerns': sheet.cell(row, 13).value,
            'dfa_non_match_reason_cnc': non_match_reason_cnc,
            'dfa_non_match_reason_si': non_match_reason_si,
            'dfa_non_match_reason_si2': non_match_reason_si2,
            'dfa_non_match_reason_pv': non_match_reason_pv,
            'dfa_odm_corrective_action_cnc': odm_corrective_action_cnc,
            'dfa_odm_corrective_action_si': odm_corrective_action_si,
            'dfa_odm_corrective_action_si2': odm_corrective_action_si2,
            'dfa_odm_corrective_action_pv': odm_corrective_action_pv,
            'dfa_pictures_from_factory': sheet.cell(row, 16).value,
            'dfa_HP_internal_discussion': sheet.cell(row, 17).value,
            'dfa_HP_internal_assessment': sheet.cell(row, 18).value,
        }
        return post_data

    # issue validate
    def issue_validate(self, post_data, product_id):
        form = DfaModelForm(data=post_data)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.dfa_product_id = product_id
            issue_list.append(post_data)
            return "pass"
        else:
            return form.errors.as_data()

    # save issues to database
    def issue_save(self, issue_list):
        for issue in issue_list:
            form = DfaModelForm(data=issue)
            instance = form.save(commit=False)
            instance.dfa_product_id = issue['dfa_product_id']
            instance.save()
        return "succeed"

    # create or update an record when dfa report import sucessfully
    def dfa_import_record_handle(self,product_name,odm_name,currentuser):
        new_item = Dfa_Report_Import_Records.objects.create(
            user= currentuser,
            import_product_name=product_name,
            import_odm= odm_name,
        )
        new_item.save()

    # log entries dfa
    def log_entries(self, request,product_name,odm_name,currentuser):
        from xadmin.models import Log
        Log.objects.create(user=currentuser,
                           ip_addr=request.META.get('REMOTE_ADDR'),
                           object_repr=str(object),
                           action_flag='',
                           message="{}-{} dfa report uploaded sucessfully".format(product_name,odm_name))

class DTReportParser():
    """
    parse DT report
    """
    def parse(self, request, workbook, currentuser):
        """ call DT method()"""
        result = self.new_template(request=request, workbook=workbook, currentuser=currentuser)
        return result

    # DT method
    def new_template(self,request, workbook, currentuser):
        global issue_list

        summary_sheet = workbook['0-Summary Data']
        odm_name = summary_sheet.cell(3, 2).value             # get odm name from summary sheet
        segment = summary_sheet.cell(4, 2).value              # get product segment from summary sheet
        product_name = summary_sheet.cell(5, 2).value         # get product name from summary sheet
        build_stage = summary_sheet.cell(7, 2).value          # get build stage from summary sheet

        sheet_names = workbook.sheetnames
        data = {}  # initial a null dict for collecting issue qty by each sheet
        issue_list = [] # initial a null list to store issues temporarily

        for sheet_name in sheet_names[4:12]:
            # count issue qty by each sheet (default is 0)
            issue_qty_by_sheet = 0
            issue_qty_gating = 0
            issue_qty_tracking = 0
            issue_qty_close = 0
            issue_qty_fix_in_next_stage = 0

            sheet = workbook[sheet_name]  # 通过sheet名称指定sheet
            rows_number = sheet.max_row  # get total row numbers
            # 调用get_picture_from_excels方法 处理sheet里面的图片，返回每张图片的具体行数和列数
            if sheet._images:
                self.result, self.buildstage, self.platform = self.get_pictures_from_excel(sheet, product_name, build_stage, sheet_name)
            else:
                self.result = {}
            # write excel contents into database by for row in range(4, rows_number):
            for row in range(5, rows_number, 4):
                if sheet.cell(row, 3).value :  # check if have contents in excel row? if yes, execute below codings; if no, break
                    unit_size = sheet.cell(row, 2).value
                    if not unit_size:
                        error = {
                            "sheet_name": sheet_name,
                            "row": row,
                            "message": "The column of unit_size is blank"  # list(result.keys())[0]
                        }
                        return [error]
                    productIds = self.create_new_product(unit_size, build_stage, odm_name, segment, currentuser)  # 调用函数 new_product 判断数据库是否已有该产品，
                    for productId in productIds:
                        issueCateory_1 = sheet.cell(row, 4).value  # write issue interaction
                        if not issueCateory_1:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message": "Issue interaction is blank"  # list(result.keys())[0]
                            }
                            return [error]
                        category_1_id = self.new_category_1(issueCateory_1)

                        issueCategory_2 = sheet.cell(row, 5).value  # write issue symptom
                        if not issueCategory_2:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message": "Issue symptom is blank"  # list(result.keys())[0]
                            }
                            return [error]
                        category_2_id = self.new_category_2(issueCategory_2, category_1_id, issueCateory_1)

                        pre_build_qty = (0 if (sheet.cell(row, 7).value) == None else (sheet.cell(row, 7).value))
                        pre_build_defcet_qty = (0 if (sheet.cell(row, 8).value) == None else (sheet.cell(row, 8).value))
                        mini1_build_qty = (0 if (sheet.cell(row + 1, 7).value) == None else (sheet.cell(row + 1, 7).value))
                        mini1_build_defcet_qty = (0 if (sheet.cell(row + 1, 8).value) == None else (sheet.cell(row + 1, 8).value))
                        mini2_build_qty = (0 if (sheet.cell(row + 2, 7).value) == None else (sheet.cell(row + 2, 7).value))
                        mini2_build_defcet_qty = (0 if (sheet.cell(row + 2, 8).value) == None else (sheet.cell(row + 2, 8).value))
                        balance_qty = (0 if (sheet.cell(row + 3, 7).value) == None else (sheet.cell(row + 3, 7).value))
                        balance_defcet_qty = (0 if (sheet.cell(row + 3, 8).value) == None else (sheet.cell(row + 3, 8).value))

                        # get and write pic path
                        if (sheet_name + str(11) + str(row)) in self.result:
                            photo = "issue/images/{}/{}/{}".format(self.platform, self.buildstage, self.result.get(sheet_name + str(11) + str(row)))
                        else:
                            photo = ""
                        # get obs
                        obs_ = sheet.cell(row, 21).value
                        if obs_ == '' or obs_ == 'N/A' or obs_ == 'N' or obs_ == 0:
                            obs = 'NA'
                        else:
                            obs = obs_
                        # get error code
                        error = sheet.cell(row, 23).value  # write issue error code number
                        if error == '' or error == "N/A" or error == 'N' or error == 0:
                            errorCode = 'NA'
                        else:
                            errorCode = error

                        post_data = {
                            "platformName_id":productId,
                            "processName":sheet_name,
                            "issue_desc":sheet.cell(row, 3).value,
                            "issue_interaction_id":category_1_id,
                            "issue_symptom_id":category_2_id,
                            "impact_scope":sheet.cell(row, 6).value,
                            "pre_build_qty":pre_build_qty,
                            "pre_build_defcet_qty": pre_build_defcet_qty,
                            "mini_build_qty": mini1_build_qty,
                            "mini_build_defcet_qty": mini1_build_defcet_qty,
                            "mini2_build_qty": mini2_build_qty,
                            "mini2_build_defcet_qty": mini2_build_defcet_qty,
                            "balance_qty": balance_qty,
                            "balance_defcet_qty": balance_defcet_qty,
                            "input_qty": pre_build_qty + mini1_build_qty + mini2_build_qty + balance_qty,
                            "defect_qty": pre_build_defcet_qty + mini1_build_defcet_qty + mini2_build_defcet_qty + balance_defcet_qty,
                            "fail_rate_stage": "NA",
                            "sn":sheet.cell(row, 10).value,
                            "sku":sheet.cell(row, 11).value,
                            "photo":photo,

                            "issue_analysis": sheet.cell(row, 13).value.replace("\n", "<br/>"),
                            "root_cause": sheet.cell(row, 14).value.replace("\n", "<br/>"),
                            "root_cause_category": sheet.cell(row, 15).value,
                            "short_term": sheet.cell(row, 16).value.replace("\n", "<br/>"),
                            "short_term_category": sheet.cell(row, 17).value,
                            "long_term": sheet.cell(row, 18).value.replace("\n", "<br/>"),
                            "long_term_category": sheet.cell(row, 19).value,
                            "status": sheet.cell(row, 20).value,
                            "obs": obs,
                            "duplicate": sheet.cell(row, 22).value ,
                            "errorCode": errorCode,
                            "repeating": sheet.cell(row, 24).value,
                            "repeatingstage": sheet.cell(row, 25).value,
                            "buildstage": sheet.cell(row, 26).value,
                            "owner": sheet.cell(row, 27).value,
                            "updatedate": sheet.cell(row, 28).value,
                            "obs_link": "http://pulsarweb.twn.hp.com/Nebula//ReportRunner/RunDetailWithObsIds?observationIds={}".format(obs),
                        }

                        #  call issue_validate to verify issue contents
                        result = self.issue_validate(post_data, productId, category_1_id, category_2_id)
                        if result == "pass":
                            # count issue quantity
                            issue_qty_by_sheet += 1
                            if sheet.cell(row, 20).value == "Gating":
                                issue_qty_gating += 1
                            elif sheet.cell(row, 20).value == "Tracking":
                                issue_qty_tracking += 1
                            elif sheet.cell(row, 20).value == "Close":
                                issue_qty_close += 1
                            elif sheet.cell(row, 20).value == "Fix in next phase":
                                issue_qty_fix_in_next_stage += 1
                            else:
                                issue_qty_tracking += 1
                        else:
                            error = {
                                "sheet_name": sheet_name,
                                "row": row,
                                "message":result  #list(result.keys())[0]
                            }
                            return [error]

            data[sheet_name] = {
                "toatl_issue_qty": issue_qty_by_sheet,
                "gating": issue_qty_gating,
                "tracking": issue_qty_tracking,
                "close": issue_qty_close,
                "fix_in_next_stage": issue_qty_fix_in_next_stage
            }

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            # import logs
            self.safelaunch_import_record(product_name, build_stage, currentuser)
            # log entries
            self.log_entries(request, product_name, build_stage, currentuser)
            return [product_name, build_stage, odm_name, data]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # issue validate
    def issue_validate(self, post_data, productId, category_1_id, category_2_id):
        form = DesktopIssueModelForm(data=post_data)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.platformName_id = productId
            instance.issue_interaction_id = category_1_id
            instance.issue_symptom_id = category_2_id

            issue_list.append(post_data)
            return "pass"
        else:
            return form.errors.as_data()

    # save issues to database
    def issue_save(self, issue_list):
        for issue in issue_list:
            form = DesktopIssueModelForm(data=issue)
            instance = form.save(commit=False)
            instance.platformName_id = issue['platformName_id']
            instance.issue_interaction_id = issue['issue_interaction_id']
            instance.issue_symptom_id = issue['issue_symptom_id']
            instance.save()
        return "succeed"

    # handel date formate because that the excel 读出数字日期（44557）转换函数
    def date_conversion(self, date):
        delta = datetime.timedelta(days=date)
        newdate = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
        return datetime.datetime.strftime(newdate, '%Y-%m-%d')

    # automatically handle product infomation
    def create_new_product(self, unit_size, build_stage, odm_name, segment, currentuser):
        """
        check if this product already in the database 使用productName查询数据库里是否已有该产品,
            if yes, then return product id
            if no, create a new one and return the product id
        """
        product_id =[] # initial a null list to record product IDs

        user = User.objects.filter(username=currentuser)
        unit_size = unit_size.split(",") # convert unit_size to list by ","
        for product in unit_size:
            obj = Products.objects.filter(ProductName=product.strip())
            if obj:
                product_id.append(obj[0].id)
                Products.objects.filter(ProductName=product).update(ProductPhase=build_stage) # update build stage only
            else:
                new_product = Products.objects.create(
                    ProductName = product,
                    Product_Segment = segment,
                    ProductPhase = build_stage,
                    PartnerName = odm_name,
                    user = currentuser,
                    PE = [(res.first_name + " " + res.last_name) for res in user][0]
                )
                new_product.save()
                product_id.append(new_product.id)
        return product_id

    # automatically handle issue's category
    def new_category_1(self, issueCategory_1):
        """
        check if this issue category already in the database
            if yes, return id
            if no, create a new category_1 and also category_2
        """
        category_1 =  SymptomCategory_First.objects.filter(category_Name=issueCategory_1)
        if category_1:
            return category_1[0].id
        else:
            newCategory = SymptomCategory_First.objects.create(
                category_Name = issueCategory_1, # write category name
                category_Desc = "This category name is {} as a 1st issue category.".format(issueCategory_1)
            )
            newCategory.save()
            return newCategory.id

    def new_category_2(self, issueCategory_2, category_1_id, issueCategory_1):
        category_2 = SymptomCategory_Second.objects.filter(category_Name=issueCategory_2)
        if category_2:
            return category_2[0].id
        else:
            newCategory = SymptomCategory_Second.objects.create(
                category_FirstType_id = category_1_id,
                category_Name=issueCategory_2,  # write category name
                category_Desc="This is a 2nd issue category belongs to {}".format(issueCategory_1)
            )
            newCategory.save()
            return newCategory.id

    # get pictures from excel in each sheet
    def get_pictures_from_excel(self, sheet, product_name, build_stage, sheet_name):
        # remove symbols from platform name
        product_name = product_name.replace(" ", "").replace("/", "_").replace("'", "").replace("\"", "").strip()
        # initial a null dict for recording pictures location(stage, sheet, row, column...)
        picture_location_dict = {}
        stages = {
            "SI": "SI",
            "PV": "PV",
            "PRB/TLD": "PRB_TLD",
            "MV (Mini1)": "MV(Mini1)",
            "MV (Mini2)": "MV(Mini2)",
            "MV (Balance Build)": "MV(Balance Build)",
            "Pre-Build (SP)": "Pre_Build(SP)",
            "Mainbuild (MVB)": "Mainbuild(MVB)",
            "CVB": "CVB",
        }
        for image in sheet._images:
            img = Image.open(BytesIO(image._data()))
            img.convert("RGB")  # fix error of "OSError: cannot write mode RGBA as JPEG" ,
            # print(image._data())
            """
            png格式图片颜色为四通道RGBA，而jpg格式是三通道RGB, （RGBA是代表Red（红色）Green（绿色）Blue（蓝色）和Alpha的色彩空间。
            虽然它有的时候被描述为一个颜色空间，但是它其实仅仅是RGB模型的附加了额外的信息。alpha通道一般用作不透明度参数。）
            """
            # img.show()
            img_name = ("{}_" + stages.get(build_stage) + "_Sheet" + sheet_name[0] + "_col{}_row{}.jpg").format(
                product_name, image.anchor._from.col, image.anchor._from.row)
            # return
            if (image.anchor._from.col, image.anchor._from.row):
                picture_location_dict["{}{}{}".format(sheet_name, str(image.anchor._from.col), str(image.anchor._from.row + 1))] = img_name

                # 单线程 -- 调用pictureUpload方法 requests.post 从本地跨域提交图片到服务器
                self.pictureUpload(product_name, stages.get(build_stage), img_name, image._data())

                # 多线程 -- 调用pictureUpload方法 requests.post 从本地跨域提交图片到服务器
                # t = threading.Thread(target=self.pictureUpload, args=(product_name, stages.get(build_stage), img_name, image._data()))
                # t.start()

        return picture_location_dict, stages.get(build_stage), product_name

    # 通过requests.post模拟游览器提交图片和信息
    def pictureUpload(self, product_name, stage_name, img_name, image):
        # url = 'http://127.0.0.1:8000/pictures_upload/'  # this url for localhost test only
        url = 'http://15.36.145.93/pictures_upload/'    # this url for server
        myData = {
            "product_name": product_name,
            "stage_name": stage_name,
            "pic_name": img_name
        }
        myFiles = {
            "img": image
        }
        res = requests.post(url, files=myFiles, data=myData)
        res.close()

    # import logs
    def safelaunch_import_record(self, prodct_name, build_stage, currentuser):
        # print("record is:", build_stage)
        from npi.models import DataImportRecords
        # 记录数据导入信息
        importRecord = DataImportRecords.objects.create(
            import_product_name=prodct_name,
            import_product_phase=build_stage,
            user=currentuser
        )
        importRecord.save()

    # write a log into Log entries after safe-launch report uploaded
    def log_entries(self, request, product_name, build_stage, currentuser):
        from xadmin.models import Log
        Log.objects.create(user=currentuser,
                           ip_addr=request.META.get('REMOTE_ADDR'),
                           object_repr=str(object),
                           action_flag='',
                           message="{}-{} safelaunch report was successfully uploaded".format(product_name, build_stage))

class IcReportParser():
    """
    parse ic tear down report
    """
    def parse(self, request, workbook, currentuser):
        """ call dfa method()"""
        result = self.ic_template(request=request, workbook=workbook, currentuser=currentuser)
        return result

    # parse ic template
    def ic_template(self,request, workbook, currentuser):
        # 取得IC EXCEL Sheet name
        ic_sheet_name = workbook.sheetnames[0]
        ic_sheet = workbook[ic_sheet_name]
        ic_Cycle = ic_sheet.cell(1, 2).value
        ic_Project_name = ic_sheet.cell(2, 2).value

        global issue_list
        issue_list = []  # initial a null list to store issues temporarily

        rows_number = ic_sheet.max_row  #get total row numbers
        empty_time = 0 #檢查最大行數之後，有三次空白，才為真正的最大行數

        for row in range(9, rows_number):
            checkingItem = ic_sheet.cell(row, 1).value
            checking_Vendor_PN = ic_sheet.cell(row, 3).value
            if checkingItem:
                post_data = self.ic_item_handle(ic_sheet, checkingItem, checking_Vendor_PN, row)
                empty_time = 0
                #  call issue_validate to verify issue contents
                result = self.issue_validate(post_data)
                if result == "pass":
                    pass
                else:
                    error = {
                        "sheet_name": ic_sheet_name,
                        "row": row,
                        "message": result  # list(result.keys())[0]
                    }
                    return [error]

            else:
                empty_time = empty_time+1
                if empty_time == 3:
                    break

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            # call method of dfa_import_record_handle to create or update an record when the above steps finished
            self.ic_import_record_handle(ic_Project_name, ic_Cycle, currentuser)
            # log entries
            self.log_entries(request, ic_Project_name, ic_Cycle, currentuser)
            return [ic_Project_name, ic_Cycle]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # handle ic item
    def ic_item_handle(self,ic_sheet,checkingItem,checking_Vendor_PN,row):

        #檢查第7欄位，有沒有合併欄位
        cell_merg = ic_sheet.cell(row=row, column=7)
        if isinstance(cell_merg, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in ic_sheet.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                if cell_merg.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell_merg = ic_sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        IC_qty_merg = cell_merg.value

        # 檢查第5欄位，替換B/S
        Part_Owner_clearn=ic_sheet.cell(row, 5).value
        if Part_Owner_clearn == 'B/S':
            Part_Owner_clearn = 'BS'

        post_data={
            'ic_Cycle': ic_sheet.cell(1, 2).value,
            'ic_Project_name': ic_sheet.cell(2, 2).value,
            'ic_HW_PM': ic_sheet.cell(3, 2).value,
            'ic_ODM': ic_sheet.cell(4, 2).value,
            'ic_Segment': ic_sheet.cell(5, 2).value,
            'ic_Chipset_Type': ic_sheet.cell(6, 2).value,
            'ic_Phase': ic_sheet.cell(7, 2).value,
            'ic_IC_Function': ic_sheet.cell(row, 1).value,
            'ic_Vendor': ic_sheet.cell(row, 2).value,
            'ic_Vendor_PN': ic_sheet.cell(row, 3).value,
            'ic_Description': ic_sheet.cell(row, 4).value,
            'ic_Part_Owner': Part_Owner_clearn,
            'ic_HP_PN': ic_sheet.cell(row, 6).value,
            'ic_IC_qty': IC_qty_merg,
            'ic_Full_Feature': ic_sheet.cell(row, 8).value,
            'ic_Defeature': ic_sheet.cell(row, 9).value,
            'ic_Vendor_2nd': ic_sheet.cell(row, 10).value,
            'ic_Vendor_PN_2nd': ic_sheet.cell(row, 11).value,
            'ic_Vendor_3rd': ic_sheet.cell(row, 12).value,
            'ic_Vendor_PN_3rd': ic_sheet.cell(row, 13).value,
            'ic_Vendor_4th': ic_sheet.cell(row, 14).value,
            'ic_Vendor_PN_4th': ic_sheet.cell(row, 15).value,
        }
        return post_data

    # issue validate
    def issue_validate(self, post_data):
        form = ICModelForm(data=post_data)
        if form.is_valid():
            instance = form.save(commit=False)
            issue_list.append(post_data)
            return "pass"
        else:
            return form.errors.as_data()

    # save issues to database
    def issue_save(self, issue_list):
        for issue in issue_list:
            form = ICModelForm(data=issue)
            instance = form.save(commit=False)
            instance.save()
        return "succeed"

    # create or update an record when dfa report import sucessfully
    def ic_import_record_handle(self,ic_Project_name, ic_Cycle,currentuser):
        new_item = IC_Report_Import_Records.objects.create(
            user= currentuser,
            import_Project_name=ic_Project_name,
            import_Cycle= ic_Cycle,
        )
        new_item.save()

    # log entries dfa
    def log_entries(self, request,ic_Project_name, ic_Cycle, currentuser):
        from xadmin.models import Log
        Log.objects.create(user=currentuser,
                           ip_addr=request.META.get('REMOTE_ADDR'),
                           object_repr=str(object),
                           action_flag='',
                           message="{}-{} IC report uploaded sucessfully".format(ic_Project_name,ic_Cycle))

class DfmReportParser():
    """
    parse Dfm tear down report
    """
    def parse(self, request, workbook, currentuser,report_version):
        """ check Dfm template version then assign the right method()"""

        if report_version == "Ver:1.63":# report_version == Ver:1.63
            result = self.oldest_template(request=request, workbook=workbook, currentuser=currentuser)
            return result

        elif report_version == "Ver:2.0":# report_version == Ver:2.0
            result = self.old_template(request=request, workbook=workbook, currentuser=currentuser)
            return result

        else: # report_version == Ver:3.0a
            result = self.new_template(request=request, workbook=workbook, currentuser=currentuser)
            return result

    # parse old template Ver:1.63
    def oldest_template(self, request, workbook, currentuser):

        dfm_sheet = workbook['DFm ']
        global issue_list
        issue_list = [] # initial a null list to store issues temporarily
        """
        以下product資料要自己手動輸入
        """
        platform_type = None
        platform_segment = None
        product_type = None
        product_name = "Warpath14LX"
        product_size = None
        odm_name = None
        build_year = None
        if build_year:#取YEAR
            build_year = int(build_year)
        rcto = None

        data = {
            'product_name': product_name,
            'platform_type': platform_type,
            'platform_segment': platform_segment,
            'product_type': product_type,
            'product_size': product_size,
            'odm_name': odm_name,
            'build_year': build_year,
            'rcto': rcto,

            'a_cover_material': None,
            'a_cover_surface': None,
            'a_cover_thickness': None,
            'a_cover_vendor1': None,
            'a_cover_vendor2': None,
            'a_cover_vendor3': None,
            'a_cover_vendor4': None,
            'a_cover_gule': None,
            'a_bonding_vendor1': None,
            'a_bonding_vendor2': None,
            'a_bonding_vendor3': None,

            'b_cover_material': None,
            'b_cover_surface': None,
            'b_cover_thickness': None,
            'b_cover_vendor1': None,
            'b_cover_vendor2': None,
            'b_cover_vendor3': None,
            'b_cover_vendor4': None,
            'b_cover_gule': None,
            'b_bonding_vendor1': None,
            'b_bonding_vendor2': None,
            'b_bonding_vendor3': None,

            'c_cover_material': None,
            'c_cover_surface': None,
            'c_cover_thickness': None,
            'c_cover_vendor1': None,
            'c_cover_vendor2': None,
            'c_cover_vendor3': None,
            'c_cover_vendor4': None,
            'c_cover_gule': None,
            'c_bonding_vendor1': None,
            'c_bonding_vendor2': None,
            'c_bonding_vendor3': None,

            'd_cover_material': None,
            'd_cover_surface': None,
            'd_cover_thickness': None,
            'd_cover_vendor1':  None,
            'd_cover_vendor2':  None,
            'd_cover_vendor3':  None,
            'd_cover_vendor4':  None,
            'd_cover_gule':  None,
            'd_bonding_vendor1':  None,
            'd_bonding_vendor2':  None,
            'd_bonding_vendor3':  None,

            'hinge_material': None,
            'hinged_cover_surface':  None,
            'hinge_thickness':  None,
            'hinge_vendor1': None,
            'hinge_vendor2':  None,
            'hinge_vendor3':  None,
            'hinge_vendor4':  None,
            'hinge_gule':  None,
            'hinge_gule_vendor1':  None,
            'hinge_gule_vendor2':  None,
            'hinge_gule_vendor3':  None,

            'touch_panel_bonding_type':  None,
            'touch_panel_vendor1':  None,
            'touch_panel_vendor2':  None,
            'touch_panel_vendor3':  None,
            'touch_panel_solution_1':  None,
            'touch_panel_solution_2':  None,
            'touch_panel_size':  None,

            'touch_pad_solution_1':  None,
            'touch_pad_solution_2':  None,
            'touch_pad_size':  None,

        }
        """
        product資料要自己手動輸入
        """
        if not product_name:
            error = {
                "The column of product_name is blank"
            }
            return [error]
        product_id = self.new_product(data,currentuser)

        rows_number = dfm_sheet.max_row  #get total row numbers
        empty_time = 0 #檢查最大行數之後，有三次空白，才為真正的最大行數

        for row in range(18, rows_number):  # rows_number = 100 写固定的行数，节省资源
            checkingItem = dfm_sheet.cell(row, 2).value #Issue Category
            if checkingItem:
                dfm_check_item = (dfm_sheet.cell(row, 2).value).replace("\n","\r\n")  # 使用.replace("\n", "\r\n")处理excel单元格里有换行问题 "\r\n"是textarea换行符
                check_item_id = self.dfm_check_item_handle_oldest(dfm_check_item, dfm_sheet.cell(row, 1).value,dfm_sheet.cell(row, 3).value, product_name)
                post_data = self.dfm_review_item_handle_oldest(dfm_sheet, check_item_id, product_id, row)
                empty_time = 0
                #  call issue_validate to verify issue contents
                result = self.issue_validate(post_data,product_id,check_item_id)
                if result == "pass":
                    pass
                else:
                    error = {
                        "sheet_name": "DFm ",
                        "row": row,
                        "message": result  # list(result.keys())[0]
                    }
                    return [error]

            else:
                empty_time = empty_time+1
                if empty_time == 3:
                    break

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            dfm_stage = "MV"
            # call method of dfa_import_record_handle to create or update an record when the above steps finished
            self.dfm_import_record_handle(product_name,dfm_stage,currentuser)
            # log entries
            self.log_entries(request, product_name, dfm_stage, currentuser)
            return [product_name,dfm_stage]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # handle checking items Ver:1.63
    def dfm_check_item_handle_oldest(self, dfm_check_item, dfm_item_item_no, dfm_item_priority, dfm_item_attributes):
            """
            check if this check item in reprot and compare with general check items
                if yes, return dfm_review_item_desc_id
                if no, create a new check item under the prodcut
            """
            item = Dfm_General_Checklist.objects.filter(dfm_item_desc=dfm_check_item,dfm_item_version="Ver:1.63")  # 通过从excel单元格获取到的item_desc，然后查询数据库获取id,priority...
            if item:
                item.update_or_create(defaults={
                    'dfm_item_attributes': 'Common',
                })
                return item[0].id
            else:
                item = Dfm_General_Checklist.objects.create(
                    dfm_item_item_no=dfm_item_item_no,
                    dfm_item_desc=dfm_check_item,
                    dfm_item_priority=dfm_item_priority,
                    dfm_item_attributes=dfm_item_attributes,
                    dfm_item_version='Ver:1.63',
                )
            item.save()
            return item.id

    # creare or update new item Ver:1.63
    def dfm_review_item_handle_oldest(self, dfm_sheet, check_item_id, product_id, row):
        nud = dfm_sheet.cell(row, 10).value
        cnc = dfm_sheet.cell(row, 4).value
        si = dfm_sheet.cell(row, 5).value
        pv = dfm_sheet.cell(row, 7).value
        mv = dfm_sheet.cell(row, 9).value

        post_data = {
            'dfm_review_item_no': dfm_sheet.cell(row, 1).value,
            'dfm_review_item_desc_id': check_item_id,
            'dfm_product_id': product_id,
            #'dfm_product_part_category': dfm_sheet.cell(row, 9).value,
            'dfm_product_issue_symptom': dfm_sheet.cell(row, 11).value,  #是Ver:1.63的Feature欄位
            'dfm_product_design_structures': dfm_sheet.cell(row, 12).value,#是Ver:1.63的Category欄位(他被隱藏起來)
            'dfm_product_nonmacth_item': dfm_sheet.cell(row, 14).value,
            'dfm_product_odm_actions': dfm_sheet.cell(row, 15).value,
            'dfm_product_solution_category': dfm_sheet.cell(row, 16).value,
            'dfm_product_nud': nud if nud else '...',
            'dfm_product_cnc': cnc if cnc else '...',
            'dfm_product_si': si if si else '...',
            'dfm_product_pv': pv if pv else '...',
            'dfm_product_mv': mv if mv else '...',
        }
        return post_data

    # parse old template Ver:2.0
    def old_template(self, request, workbook, currentuser):

        dfm_sheet = workbook['DFm ']
        global issue_list
        issue_list = [] # initial a null list to store issues temporarily

        platform_type = dfm_sheet.cell(3, 13).value
        platform_segment = dfm_sheet.cell(4, 13).value
        product_type = dfm_sheet.cell(5, 13).value
        product_name = dfm_sheet.cell(6, 13).value
        product_size = dfm_sheet.cell(7, 13).value
        odm_name = dfm_sheet.cell(8, 13).value
        build_year = dfm_sheet.cell(9, 13).value
        if build_year:#取YEAR
            build_year=str(build_year)
            build_year_split = build_year.split('-')
            build_year = build_year_split[0]
            build_year = int(build_year)
        rcto = dfm_sheet.cell(3, 15).value

        key_materials = workbook['Key Material information']
        data = {
            'product_name': product_name,
            'platform_type': platform_type,
            'platform_segment': platform_segment,
            'product_type': product_type,
            'product_size': product_size,
            'odm_name': odm_name,
            'build_year': build_year,
            'rcto': rcto,

            'a_cover_material': key_materials.cell(4, 3).value,
            'a_cover_surface': key_materials.cell(4, 4).value,
            'a_cover_thickness': key_materials.cell(4, 5).value,
            'a_cover_vendor1': key_materials.cell(4, 6).value,
            'a_cover_vendor2': key_materials.cell(4, 7).value,
            'a_cover_vendor3': key_materials.cell(4, 8).value,
            'a_cover_vendor4': key_materials.cell(4, 9).value,
            'a_cover_gule': key_materials.cell(12, 3).value,
            'a_bonding_vendor1': key_materials.cell(12, 4).value,
            'a_bonding_vendor2': key_materials.cell(12, 5).value,
            'a_bonding_vendor3': key_materials.cell(12, 6).value,

            'b_cover_material': key_materials.cell(5, 3).value,
            'b_cover_surface': key_materials.cell(5, 4).value,
            'b_cover_thickness': key_materials.cell(5, 5).value,
            'b_cover_vendor1': key_materials.cell(5, 6).value,
            'b_cover_vendor2': key_materials.cell(5, 7).value,
            'b_cover_vendor3': key_materials.cell(5, 8).value,
            'b_cover_vendor4': key_materials.cell(5, 9).value,
            'b_cover_gule': key_materials.cell(13, 3).value,
            'b_bonding_vendor1': key_materials.cell(13, 4).value,
            'b_bonding_vendor2': key_materials.cell(13, 5).value,
            'b_bonding_vendor3': key_materials.cell(13, 6).value,

            'c_cover_material': key_materials.cell(6, 3).value,
            'c_cover_surface': key_materials.cell(6, 4).value,
            'c_cover_thickness': key_materials.cell(6, 5).value,
            'c_cover_vendor1': key_materials.cell(6, 6).value,
            'c_cover_vendor2': key_materials.cell(6, 7).value,
            'c_cover_vendor3': key_materials.cell(6, 8).value,
            'c_cover_vendor4': key_materials.cell(6, 9).value,
            'c_cover_gule': key_materials.cell(14, 3).value,
            'c_bonding_vendor1': key_materials.cell(14, 4).value,
            'c_bonding_vendor2': key_materials.cell(14, 5).value,
            'c_bonding_vendor3': key_materials.cell(14, 6).value,

            'd_cover_material': key_materials.cell(7, 3).value,
            'd_cover_surface': key_materials.cell(7, 4).value,
            'd_cover_thickness': key_materials.cell(7, 5).value,
            'd_cover_vendor1': key_materials.cell(7, 6).value,
            'd_cover_vendor2': key_materials.cell(7, 7).value,
            'd_cover_vendor3': key_materials.cell(7, 8).value,
            'd_cover_vendor4': key_materials.cell(7, 9).value,
            'd_cover_gule': key_materials.cell(15, 3).value,
            'd_bonding_vendor1': key_materials.cell(15, 4).value,
            'd_bonding_vendor2': key_materials.cell(15, 5).value,
            'd_bonding_vendor3': key_materials.cell(15, 6).value,

            'hinge_material': key_materials.cell(8, 3).value,
            'hinged_cover_surface': key_materials.cell(8, 4).value,
            'hinge_thickness': key_materials.cell(8, 5).value,
            'hinge_vendor1': key_materials.cell(8, 6).value,
            'hinge_vendor2': key_materials.cell(8, 7).value,
            'hinge_vendor3': key_materials.cell(8, 8).value,
            'hinge_vendor4': key_materials.cell(8, 9).value,
            'hinge_gule': key_materials.cell(16, 3).value,
            'hinge_gule_vendor1': key_materials.cell(16, 4).value,
            'hinge_gule_vendor2': key_materials.cell(16, 5).value,
            'hinge_gule_vendor3': key_materials.cell(16, 6).value,

            'touch_panel_bonding_type': key_materials.cell(20, 3).value,
            'touch_panel_vendor1': key_materials.cell(20, 4).value,
            'touch_panel_vendor2': key_materials.cell(20, 5).value,
            'touch_panel_vendor3': key_materials.cell(20, 6).value,
            'touch_panel_solution_1': key_materials.cell(24, 3).value,
            'touch_panel_solution_2': key_materials.cell(24, 4).value,
            'touch_panel_size': key_materials.cell(24, 5).value,

            'touch_pad_solution_1': key_materials.cell(25, 3).value,
            'touch_pad_solution_2': key_materials.cell(25, 4).value,
            'touch_pad_size': key_materials.cell(25, 5).value,

        }
        if not product_name:
            error = {
                "The column of product_name is blank"
            }
            return [error]
        product_id = self.new_product(data, currentuser)

        rows_number = dfm_sheet.max_row  #get total row numbers
        empty_time = 0 #檢查最大行數之後，有三次空白，才為真正的最大行數

        for row in range(31, rows_number):  # rows_number = 100 写固定的行数，节省资源
            checkingItem = dfm_sheet.cell(row, 2).value #Issue Category
            if checkingItem:
                dfm_check_item = (dfm_sheet.cell(row, 2).value).replace("\n","\r\n")  # 使用.replace("\n", "\r\n")处理excel单元格里有换行问题 "\r\n"是textarea换行符
                check_item_id = self.dfm_check_item_handle(dfm_check_item, dfm_sheet.cell(row, 1).value,dfm_sheet.cell(row, 3).value, product_name)
                post_data = self.dfm_review_item_handle(dfm_sheet, check_item_id, product_id, row)
                empty_time = 0
                #  call issue_validate to verify issue contents
                result = self.issue_validate(post_data, product_id, check_item_id)
                if result == "pass":
                    pass
                else:
                    error = {
                        "sheet_name": "DFm ",
                        "row": row,
                        "message": result  # list(result.keys())[0]
                    }
                    return [error]
            else:
                empty_time = empty_time+1
                if empty_time == 3:
                    break

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            dfm_stage = ""
            if dfm_sheet.cell(32, 7).value:  # mv
                dfm_stage = "MV"
            elif dfm_sheet.cell(32, 6).value:  # pv
                dfm_stage = "PV"
            elif dfm_sheet.cell(32, 5).value:  # si
                dfm_stage = "SI"
            else:  # cnc
                dfm_stage = "CNC"

            # call method of dfa_import_record_handle to create or update an record when the above steps finished
            self.dfm_import_record_handle(product_name, dfm_stage, currentuser)
            # log entries
            self.log_entries(request, product_name, dfm_stage, currentuser)
            return [product_name, dfm_stage, odm_name]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # handle checking items Ver:2.0
    def dfm_check_item_handle(self, dfm_check_item, dfm_item_item_no, dfm_item_priority, dfm_item_attributes):
            """
            check if this check item in reprot and compare with general check items
                if yes, return dfm_review_item_desc_id
                if no, create a new check item under the prodcut
            """
            item = Dfm_General_Checklist.objects.filter(dfm_item_desc=dfm_check_item,dfm_item_version="Ver:2.0")  # 通过从excel单元格获取到的item_desc，然后查询数据库获取id,priority...
            if item:
                item.update_or_create(defaults={
                    'dfm_item_attributes': 'Common',
                })
                return item[0].id
            else:
                item = Dfm_General_Checklist.objects.create(
                    dfm_item_item_no=dfm_item_item_no,
                    dfm_item_desc=dfm_check_item,
                    dfm_item_priority=dfm_item_priority,
                    dfm_item_attributes=dfm_item_attributes,
                )
            item.save()
            return item.id

    # creare or update new item Ver:2.0
    def dfm_review_item_handle(self, dfm_sheet, check_item_id, product_id, row):
            cnc = dfm_sheet.cell(row, 4).value
            si = dfm_sheet.cell(row, 5).value
            pv = dfm_sheet.cell(row, 6).value
            mv = dfm_sheet.cell(row, 7).value
            nud = dfm_sheet.cell(row, 8).value

            post_data ={
                'dfm_review_item_no': dfm_sheet.cell(row, 1).value,
                'dfm_review_item_desc_id': check_item_id,
                'dfm_product_id': product_id,
                'dfm_product_part_category': dfm_sheet.cell(row, 9).value,
                'dfm_product_issue_symptom': dfm_sheet.cell(row, 10).value,
                'dfm_product_design_structures': dfm_sheet.cell(row, 11).value,
                'dfm_product_nonmacth_item': dfm_sheet.cell(row, 13).value,
                'dfm_product_odm_actions': dfm_sheet.cell(row, 14).value,
                'dfm_product_solution_category': dfm_sheet.cell(row, 15).value,
                'dfm_product_nud': nud if nud else '...',
                'dfm_product_cnc': cnc if cnc else '...',
                'dfm_product_si': si if si else '...',
                'dfm_product_pv': pv if pv else '...',
                'dfm_product_mv': mv if mv else '...',
            }
            return post_data

    # parse new template Ver:3.0a
    def new_template(self,request, workbook, currentuser):

        dfm_sheet = workbook['DFm ']
        global issue_list
        issue_list = [] # initial a null list to store issues temporarily

        platform_type = dfm_sheet.cell(3, 13).value
        platform_segment = dfm_sheet.cell(4, 13).value
        product_type = dfm_sheet.cell(5, 13).value
        product_name = dfm_sheet.cell(6, 13).value
        product_size = dfm_sheet.cell(7, 13).value
        odm_name = dfm_sheet.cell(8, 13).value
        build_year = dfm_sheet.cell(9, 13).value
        if build_year:#取YEAR
            build_year=str(build_year)
            build_year_split = build_year.split('-')
            build_year = build_year_split[0]
            build_year = int(build_year)
        rcto = dfm_sheet.cell(3, 15).value

        key_materials = workbook['Key Material information']
        data = {
            'product_name': product_name,
            'platform_type': platform_type,
            'platform_segment': platform_segment,
            'product_type': product_type,
            'product_size': product_size,
            'odm_name': odm_name,
            'build_year': build_year,
            'rcto': rcto,

            'a_cover_material': key_materials.cell(4, 3).value,
            'a_cover_surface': key_materials.cell(4, 4).value,
            'a_cover_thickness': key_materials.cell(4, 5).value,
            'a_cover_vendor1': key_materials.cell(4, 6).value,
            'a_cover_vendor2': key_materials.cell(4, 7).value,
            'a_cover_vendor3': key_materials.cell(4, 8).value,
            'a_cover_vendor4': key_materials.cell(4, 9).value,
            'a_cover_gule': key_materials.cell(12, 3).value,
            'a_bonding_vendor1': key_materials.cell(12, 4).value,
            'a_bonding_vendor2': key_materials.cell(12, 5).value,
            'a_bonding_vendor3': key_materials.cell(12, 6).value,

            'b_cover_material': key_materials.cell(5, 3).value,
            'b_cover_surface': key_materials.cell(5, 4).value,
            'b_cover_thickness': key_materials.cell(5, 5).value,
            'b_cover_vendor1': key_materials.cell(5, 6).value,
            'b_cover_vendor2': key_materials.cell(5, 7).value,
            'b_cover_vendor3': key_materials.cell(5, 8).value,
            'b_cover_vendor4': key_materials.cell(5, 9).value,
            'b_cover_gule': key_materials.cell(13, 3).value,
            'b_bonding_vendor1': key_materials.cell(13, 4).value,
            'b_bonding_vendor2': key_materials.cell(13, 5).value,
            'b_bonding_vendor3': key_materials.cell(13, 6).value,

            'c_cover_material': key_materials.cell(6, 3).value,
            'c_cover_surface': key_materials.cell(6, 4).value,
            'c_cover_thickness': key_materials.cell(6, 5).value,
            'c_cover_vendor1': key_materials.cell(6, 6).value,
            'c_cover_vendor2': key_materials.cell(6, 7).value,
            'c_cover_vendor3': key_materials.cell(6, 8).value,
            'c_cover_vendor4': key_materials.cell(6, 9).value,
            'c_cover_gule': key_materials.cell(14, 3).value,
            'c_bonding_vendor1': key_materials.cell(14, 4).value,
            'c_bonding_vendor2': key_materials.cell(14, 5).value,
            'c_bonding_vendor3': key_materials.cell(14, 6).value,

            'd_cover_material': key_materials.cell(7, 3).value,
            'd_cover_surface': key_materials.cell(7, 4).value,
            'd_cover_thickness': key_materials.cell(7, 5).value,
            'd_cover_vendor1': key_materials.cell(7, 6).value,
            'd_cover_vendor2': key_materials.cell(7, 7).value,
            'd_cover_vendor3': key_materials.cell(7, 8).value,
            'd_cover_vendor4': key_materials.cell(7, 9).value,
            'd_cover_gule': key_materials.cell(15, 3).value,
            'd_bonding_vendor1': key_materials.cell(15, 4).value,
            'd_bonding_vendor2': key_materials.cell(15, 5).value,
            'd_bonding_vendor3': key_materials.cell(15, 6).value,

            'hinge_material': key_materials.cell(8, 3).value,
            'hinged_cover_surface': key_materials.cell(8, 4).value,
            'hinge_thickness': key_materials.cell(8, 5).value,
            'hinge_vendor1': key_materials.cell(8, 6).value,
            'hinge_vendor2': key_materials.cell(8, 7).value,
            'hinge_vendor3': key_materials.cell(8, 8).value,
            'hinge_vendor4': key_materials.cell(8, 9).value,
            'hinge_gule': key_materials.cell(16, 3).value,
            'hinge_gule_vendor1': key_materials.cell(16, 4).value,
            'hinge_gule_vendor2': key_materials.cell(16, 5).value,
            'hinge_gule_vendor3': key_materials.cell(16, 6).value,

            'touch_panel_bonding_type': key_materials.cell(20, 4).value,
            'touch_panel_vendor1': key_materials.cell(20, 5).value,
            'touch_panel_vendor2': key_materials.cell(20, 6).value,
            'touch_panel_vendor3': key_materials.cell(20, 7).value,
            'touch_panel_solution_1': key_materials.cell(20, 3).value,
            'touch_panel_solution_2': '',
            'touch_panel_size': '',

            'touch_pad_solution_1': key_materials.cell(24, 3).value,
            'touch_pad_solution_2': key_materials.cell(24, 4).value,
            'touch_pad_size': key_materials.cell(24, 5).value,

        }
        if not product_name:
            error = {
                "The column of product_name is blank"
            }
            return [error]
        product_id = self.new_product(data, currentuser)

        # create = Dfm_Review_Result()
        rows_number = dfm_sheet.max_row  # get total row numbers
        empty_time = 0  # 檢查最大行數之後，有三次空白，才為真正的最大行數
        for row in range(32, rows_number):  # rows_number = 100 写固定的行数，节省资源
            checkingItem = dfm_sheet.cell(row, 3).value
            if checkingItem:
                dfm_check_item = (dfm_sheet.cell(row, 3).value).replace("\n","\r\n")  # 使用.replace("\n", "\r\n")处理excel单元格里有换行问题 "\r\n"是textarea换行符
                check_item_id = self.dfm_check_item_handle_newVersion(dfm_check_item, dfm_sheet.cell(row, 1).value,dfm_sheet.cell(row, 4).value, product_name, row,dfm_sheet)
                post_data = self.dfm_review_item_handle_newVersion(dfm_sheet, check_item_id, product_id, row)
                empty_time = 0
                #  call issue_validate to verify issue contents
                result = self.issue_validate(post_data, product_id, check_item_id)
                if result == "pass":
                    pass
                else:
                    error = {
                        "sheet_name": "DFm ",
                        "row": row,
                        "message": result  # list(result.keys())[0]
                    }
                    return [error]
            else:
                empty_time = empty_time+1
                if empty_time == 3:
                    break

        # call issue_save to save issues to database
        flag = self.issue_save(issue_list)
        if flag == "succeed":
            dfm_stage = ""
            if dfm_sheet.cell(32, 8).value:  # mv
                dfm_stage = "MV"
            elif dfm_sheet.cell(32, 7).value:  # pv
                dfm_stage = "PV"
            elif dfm_sheet.cell(32, 6).value:  # si
                dfm_stage = "SI"
            else:  # cnc
                dfm_stage = "CNC"

            # call method of dfa_import_record_handle to create or update an record when the above steps finished
            self.dfm_import_record_handle(product_name, dfm_stage, currentuser)
            # log entries
            self.log_entries(request, product_name, dfm_stage, currentuser)
            return [product_name,dfm_stage, odm_name]
        else:
            error = {
                "data saving failed."
            }
            return [error]

    # handle checking items Ver:3.0a
    def dfm_check_item_handle_newVersion(self, dfm_check_item, dfm_item_item_no, dfm_item_priority, dfm_item_attributes,row,dfm_sheet):
        """
        check if this check item in reprot and compare with general check items
            if yes, return dfm_review_item_desc_id
            if no, create a new check item under the prodcut
        """
        #檢查第2欄位，有沒有合併欄位
        cell_merg = dfm_sheet.cell(row=row, column=2)
        if isinstance(cell_merg, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in dfm_sheet.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                if cell_merg.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell_merg = dfm_sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        dfm_assembly_level = cell_merg.value
        dfm_assembly_level = dfm_assembly_level.rstrip()
        dfm_assembly_level = dfm_assembly_level.lstrip()

        item = Dfm_General_Checklist.objects.filter(dfm_item_desc=dfm_check_item,dfm_item_version="Ver:3.0a")  # 通过从excel单元格获取到的item_desc，然后查询数据库获取id,priority...
        if item:
            item.update_or_create(defaults={
                'dfm_item_attributes': 'Common',
            })
            return item[0].id
        else:
            item = Dfm_General_Checklist.objects.create(
            dfm_item_item_no = dfm_item_item_no,
            dfm_assembly_level = dfm_assembly_level, # for the version 2023 +
            dfm_item_desc = dfm_check_item,
            dfm_item_priority = dfm_item_priority,
            dfm_item_attributes = dfm_item_attributes,
            dfm_item_version = 'Ver:3.0a',
        )
        item.save()
        return item.id

    # creare or update new item Ver:3.0a
    def dfm_review_item_handle_newVersion(self, dfm_sheet, check_item_id, product_id, row):
        nud = dfm_sheet.cell(row, 9).value
        cnc = dfm_sheet.cell(row, 5).value
        si = dfm_sheet.cell(row, 6).value
        pv = dfm_sheet.cell(row, 7).value
        mv = dfm_sheet.cell(row, 8).value

        post_data={
            'dfm_review_item_no': dfm_sheet.cell(row, 1).value,
            'dfm_review_item_desc_id': check_item_id,
            'dfm_product_id': product_id,
            #'dfm_product_part_category': dfm_sheet.cell(row, 9).value,
            'dfm_product_issue_symptom': dfm_sheet.cell(row, 10).value,
            'dfm_product_design_structures': dfm_sheet.cell(row, 11).value,
            'dfm_product_nonmacth_item': dfm_sheet.cell(row, 13).value,
            'dfm_product_odm_actions': dfm_sheet.cell(row, 14).value,
            'dfm_product_solution_category': dfm_sheet.cell(row, 15).value,
            'dfm_product_nud': nud if nud else '...',
            'dfm_product_cnc': cnc if cnc else '...',
            'dfm_product_si': si if si else '...',
            'dfm_product_pv': pv if pv else '...',
            'dfm_product_mv': mv if mv else '...',
        }
        return post_data

    # create or update product Ver:1.63、Ver:2.0、Ver:3.0a
    def new_product(self,data,currentuser):
        """
        check if this product already in the database 使用productName查询数据库里是否已有该产品,
            if yes, then return product id
            if no, create a new one and return the product id
        """
        item = Products.objects.filter(ProductName=data['product_name']).update_or_create(defaults={
            'ProductName': data['product_name'],
            'Platform_Type': data['platform_type'] if data['platform_type'] else '',
            'Product_Segment': data['platform_segment'] if data['platform_segment'] else '',
            'Product_Type': data['product_type'] if data['product_type'] else '',
            'Product_Size': data['product_size'],
            'PartnerName': data['odm_name'],
            "Product_RCTO_Type": data['rcto'] if data['rcto'] else '',
            "Product_MV_date" : data['build_year'],

            'Product_A_Cover_material_Type': data['a_cover_material'],
            'Product_A_Cover_material_Surface': data['a_cover_surface'],
            'Product_A_Cover_material_Thickness': data['a_cover_thickness'],
            'Product_A_Cover_material_Supplier01': data['a_cover_vendor1'],
            'Product_A_Cover_material_Supplier02': data['a_cover_vendor2'],
            'Product_A_Cover_material_Supplier03': data['a_cover_vendor3'],
            'Product_A_Cover_material_Supplier04': data['a_cover_vendor4'],
            'Product_A_Cover_Glue_Vendor': data['a_cover_gule'],
            'Product_A_Cover_Bonding_Vendor01': data['a_bonding_vendor1'],
            'Product_A_Cover_Bonding_Vendor02': data['a_bonding_vendor2'],
            'Product_A_Cover_Bonding_Vendor03': data['a_bonding_vendor3'],

            'Product_B_Cover_material_Type': data['b_cover_material'],
            'Product_B_Cover_material_Surface': data['b_cover_surface'],
            'Product_B_Cover_material_Thickness': data['b_cover_thickness'],
            'Product_B_Cover_material_Supplier01': data['b_cover_vendor1'],
            'Product_B_Cover_material_Supplier02': data['b_cover_vendor2'],
            'Product_B_Cover_material_Supplier03': data['b_cover_vendor3'],
            'Product_B_Cover_material_Supplier04': data['b_cover_vendor4'],
            'Product_B_Cover_Glue_Vendor': data['b_cover_gule'],
            'Product_B_Cover_Bonding_Vendor01': data['b_bonding_vendor1'],
            'Product_B_Cover_Bonding_Vendor02': data['b_bonding_vendor2'],
            'Product_B_Cover_Bonding_Vendor03': data['b_bonding_vendor3'],

            'Product_C_Cover_material_Type': data['c_cover_material'],
            'Product_C_Cover_material_Surface': data['c_cover_surface'],
            'Product_C_Cover_material_Thickness': data['c_cover_thickness'],
            'Product_C_Cover_material_Supplier01': data['c_cover_vendor1'],
            'Product_C_Cover_material_Supplier02': data['c_cover_vendor2'],
            'Product_C_Cover_material_Supplier03': data['c_cover_vendor3'],
            'Product_C_Cover_material_Supplier04': data['c_cover_vendor4'],
            'Product_C_Cover_Glue_Vendor': data['c_cover_gule'],
            'Product_C_Cover_Bonding_Vendor01': data['c_bonding_vendor1'],
            'Product_C_Cover_Bonding_Vendor02': data['c_bonding_vendor2'],
            'Product_C_Cover_Bonding_Vendor03': data['c_bonding_vendor3'],

            'Product_D_Cover_material_Type': data['d_cover_material'],
            'Product_D_Cover_material_Surface': data['d_cover_surface'],
            'Product_D_Cover_material_Thickness': data['d_cover_thickness'],
            'Product_D_Cover_material_Supplier01': data['d_cover_vendor1'],
            'Product_D_Cover_material_Supplier02': data['d_cover_vendor2'],
            'Product_D_Cover_material_Supplier03': data['d_cover_vendor3'],
            'Product_D_Cover_material_Supplier04': data['d_cover_vendor4'],
            'Product_D_Cover_Glue_Vendor': data['d_cover_gule'],
            'Product_D_Cover_Bonding_Vendor01': data['d_bonding_vendor1'],
            'Product_D_Cover_Bonding_Vendor02': data['d_bonding_vendor2'],
            'Product_D_Cover_Bonding_Vendor03': data['d_bonding_vendor3'],

            'Product_Hinge_material_Type': data['hinge_material'],
            'Product_Hinge_material_Surface': data['hinged_cover_surface'],
            'Product_Hinge_material_Thickness': data['hinge_thickness'],
            'Product_Hinge_material_Supplier01': data['hinge_vendor1'],
            'Product_Hinge_material_Supplier02': data['hinge_vendor2'],
            'Product_Hinge_material_Supplier03': data['hinge_vendor3'],
            'Product_Hinge_material_Supplier04': data['hinge_vendor4'],
            'Product_Hinge_Glue_Vendor': data['hinge_gule'],
            'Product_Hinge_Bonding_Vendor01': data['hinge_gule_vendor1'],
            'Product_Hinge_Bonding_Vendor02': data['hinge_gule_vendor2'],
            'Product_Hinge_Bonding_Vendor03': data['hinge_gule_vendor3'],

            'Product_TouchPanel': data['touch_panel_bonding_type'],
            'Product_TouchPanel_BondingVendor01': data['touch_panel_vendor1'],
            'Product_TouchPanel_BondingVendor02': data['touch_panel_vendor2'],
            'Product_TouchPanel_BondingVendor03': data['touch_panel_vendor3'],
            'Product_TouchPanel_Solution01': data['touch_panel_solution_1'],
            'Product_TouchPanel_Solution02': data['touch_panel_solution_2'],
            'Product_TouchPanel_Outsize': data['touch_panel_size'],

            'Product_Touchpad_Solution01': data['touch_pad_solution_1'],
            'Product_Touchpad_Solution02': data['touch_pad_solution_2'],
            'Product_Touchpad_Outsize': data['touch_pad_size'],
            'user': currentuser,
        })
        print('hey,before return product id')
        return item[0].id

    # issue validate Ver:1.63、Ver:2.0、Ver:3.0a
    def issue_validate(self,post_data,product_id,check_item_id):
        form = DfmIssueModelForm(data=post_data)
        if form.is_valid():
            instance_issue = form.save(commit=False)
            instance_issue.dfm_review_item_desc_id = check_item_id
            instance_issue.dfm_product_id = product_id
            issue_list.append(post_data)
            return "pass"
        else:
            return form.errors.as_data()

    # save issues to database Ver:1.63、Ver:2.0、Ver:3.0a
    def issue_save(self,issue_list):
        for issue in issue_list:
            form_issue = DfmIssueModelForm(data=issue)
            instance_issue = form_issue.save(commit=False)
            instance_issue.dfm_review_item_desc_id = issue['dfm_review_item_desc_id']
            instance_issue.dfm_product_id = issue['dfm_product_id']
            instance_issue.save()
        return "succeed"

    # create or update an record when dfm report import sucessfully Ver:1.63、Ver:2.0、Ver:3.0a
    def dfm_import_record_handle(self,product_name,dfm_stage,currentuser):
        new_item = DFM_Report_Import_Records.objects.filter(import_product_name=product_name)
        new_item.update_or_create(defaults={
            'user': currentuser,
            'import_product_name': product_name,
        })
        if dfm_stage == "MV":
            new_item.update_or_create(defaults={
                'import_stage_mv': 'Y'
            })
        elif dfm_stage == "PV":
            new_item.update_or_create(defaults={
                'import_stage_pv': 'Y',
            })
        elif dfm_stage == "SI":
            new_item.update_or_create(defaults={
                'import_stage_si': 'Y',
            })
        elif dfm_stage == "CNC":
            new_item.update_or_create(defaults={
                'import_stage_cnc': 'Y',
            })

    # log entries Ver:1.63、Ver:2.0、Ver:3.0a
    def log_entries(self, request, product_name, dfm_stage, currentuser):
        from xadmin.models import Log
        Log.objects.create(user=currentuser,
                           ip_addr=request.META.get('REMOTE_ADDR'),
                           object_repr=str(object),
                           action_flag='',
                           message="{}-{} dfm report was successfully uploaded".format(product_name,dfm_stage))
