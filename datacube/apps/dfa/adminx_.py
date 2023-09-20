from xadmin.filters import MultiSelectFieldListFilter
from django.contrib import messages
from django.http import HttpResponseRedirect

import xadmin
from .models import Dfa_Review_Result,Dfa_Report_Import_Records
from xadmin.layout import Main, Fieldset, Side, Row
from import_export import resources
from utils.notification import Notification

# define dfm user interface
class DfaAdmin(object):

    list_display = [
        "dfa_product",
        "dfa_object",
        "dfa_category",
        "dfa_production_line",
        "dfa_cnc",
        "dfa_si_pv",
        "dfa_mv_mp",
        "dfa_HP_commodities_design",
        "dfa_odm_product_designs",
        "dfa_odm_packing_designs_kitting_box_design",
        ]
    list_filter = [
        ("dfa_product", MultiSelectFieldListFilter),
        ("dfa_object", MultiSelectFieldListFilter),
        ("dfa_category", MultiSelectFieldListFilter),
        ("dfa_production_line", MultiSelectFieldListFilter),
        ("dfa_cnc", MultiSelectFieldListFilter),
        ("dfa_si_pv", MultiSelectFieldListFilter),
        ("dfa_mv_mp", MultiSelectFieldListFilter),
        ("dfa_non_match_reason_cnc", MultiSelectFieldListFilter),
        ("dfa_non_match_reason_si", MultiSelectFieldListFilter),
        ("dfa_non_match_reason_si2", MultiSelectFieldListFilter),
        ("dfa_non_match_reason_pv", MultiSelectFieldListFilter),
        ("dfa_odm_corrective_action_cnc", MultiSelectFieldListFilter),
        ("dfa_odm_corrective_action_si", MultiSelectFieldListFilter),
        ("dfa_odm_corrective_action_si2", MultiSelectFieldListFilter),
        ("dfa_odm_corrective_action_pv", MultiSelectFieldListFilter),
        ("dfa_HP_internal_discussion", MultiSelectFieldListFilter),
        ("dfa_HP_internal_assessment", MultiSelectFieldListFilter),
        ]
    search_fields = [
        "dfa_product__ProductName",
        "dfa_category",
        "dfa_object",
        "dfa_estimated_hc",
        "dfa_production_line",
        "dfa_cnc",
        "dfa_si_pv",
        "dfa_mv_mp",
        "dfa_HP_commodities_design",
        "dfa_odm_product_designs",
        "dfa_odm_packing_designs_kitting_box_design",
        "dfa_HP_packing_designs",
        "dfa_equipment_facilites_UPH_concerns",
        "dfa_non_match_reason_cnc",
        "dfa_non_match_reason_si",
        "dfa_non_match_reason_si2",
        "dfa_non_match_reason_pv",
        "dfa_odm_corrective_action_cnc",
        "dfa_odm_corrective_action_si",
        "dfa_odm_corrective_action_si2",
        "dfa_odm_corrective_action_pv",
        "dfa_pictures_from_factory",
        "dfa_HP_internal_discussion",
        "dfa_HP_internal_assessment",
    ]
    list_display_links = ["dfa_object"]
    list_per_page = 50
    date_hierarchy = 'dfa_crate_date'
    list_editable = ('dfa_object')
    model_icon = 'fa fa-cogs' #fa fa-file-text-o

    ### Dfa_review_product 注册导入导出功能
    class Issue_import_export_Resource(resources.ModelResource):
        class Meta:
            model = Dfa_Review_Result

    import_export_args = {'export_resource_class': Issue_import_export_Resource}
    import_excel = True

    # rewrite post method for getting excel DFM reprot
    def post(self, request, *args, **kwargs):
        currentuser = self.request.user  # get current user

        if 'excel' in request.FILES:
            dfa_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(dfa_report, data_only=True, keep_vba=False, keep_links=False)

            """ check Dfa template then assign the right method()"""
            dfa_sheet = workbook['Dashboard Page']
            product_name = dfa_sheet.cell(2, 13).value
            odm_name = dfa_sheet.cell(2, 7).value
            # check if the product is the first upload or repeat to upload(don't support repeated upload)
            record = Dfa_Report_Import_Records.objects.filter(import_product_name=product_name, import_odm=odm_name)

            if record:  # repeated uploads

                # raise a message on webpage
                messages.error(request,
                               "Failed to import data! error message: [ProductId:{} -- {} -- {}'dfa data already imported, repeated uploads do not supported]".format(
                                   record[0].id, record[0].import_product_name, record[0].import_odm))
                """

                # send an alert by email
                contents = "Failed to import data!  error message: [ProductId:{} -- {} -- {}'dfa data already imported, repeated uploads do not supported]".format(
                    record[0].id, record[0].import_product_name, record[0].import_odm)
                Notification().send_by_email(
                    [str(currentuser)],
                    currentuser,
                    contents,
                    record[0].import_product_name,  # platform name
                    record[0].import_odm,  # odm
                    "dfa report upload Failed",  # email title
                    None,
                    None, )
                """

            else:  # upload field
                from utils.excel_report_parser import DfaReportParser
                result = DfaReportParser().parse(request, workbook, currentuser)
                # raise a message on webpage
                messages.success(request, "{}-{}'s dfa data was imported successfully.".format(result[0], result[1]))
                """
                #   send an alert by email
                contents = "{}-{} dfm data was imported successfully.".format(result[0],result[1])
                Notification().send_by_email(
                                            [str(currentuser)],
                                            currentuser,
                                            contents,
                                            result[0],                                  # platform name
                                            result[1],                                  # odm
                                            "Dfa report uploaded Successfully",  # email title
                                            None,                                  # 返回data
                                            None,)                                 # odm name
                """
            workbook.close()
            return HttpResponseRedirect('/scpe/dfa/dfa_review_result/')
        return super(DfaAdmin, self).post(request, *args, **kwargs)


    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         'dfa_object',
                         Row('dfa_category', 'dfa_product'),
                         Row('dfa_estimated_hc', 'dfa_production_line'),
                         Row('dfa_cnc','dfa_si_pv'),
                         'dfa_mv_mp',
                         'dfa_crate_date',
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('dfa_non_match_reason_cnc','dfa_non_match_reason_si'),
                         Row('dfa_non_match_reason_si2','dfa_non_match_reason_pv'),
                         Row('dfa_odm_corrective_action_cnc', 'dfa_odm_corrective_action_si'),
                         Row('dfa_odm_corrective_action_si2', 'dfa_odm_corrective_action_pv'),
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('dfa_pictures_from_factory'),
                         Row('dfa_HP_internal_discussion', 'dfa_HP_internal_assessment'),
                         css_class='unsort no_title'
                         ),
                ),
            Side(
                Fieldset('',
                         'dfa_HP_commodities_design','dfa_odm_product_designs','dfa_odm_packing_designs_kitting_box_design','dfa_HP_packing_designs','dfa_equipment_facilites_UPH_concerns',
                         css_class='unsort no_title'
                         ),
            ),
            )
        return super(DfaAdmin, self).get_form_layout()

class Dfa_Upload_Records_Admin(object):
    list_display = ["user","import_product_name","import_odm","import_date"]
    list_display_links =["import_product_name"]

    def save_models(self):
        self.new_obj.user = self.request.user
        return super().save_models()


#后台管理项目注册
xadmin.site.register(Dfa_Review_Result, DfaAdmin)
xadmin.site.register(Dfa_Report_Import_Records, Dfa_Upload_Records_Admin)