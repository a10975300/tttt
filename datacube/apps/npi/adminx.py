import xadmin
from django.contrib import messages
from django.http import HttpResponseRedirect
from xadmin.filters import MultiSelectFieldListFilter
from .models import DataImportRecords, SymptomCategory_First,Issue,RegionalCase,DesktopIssue,SymptomCategory_Second
from xadmin.layout import Main, Fieldset, Side, Row
from product.models import Products
import warnings
warnings.simplefilter(action='ignore', category=UserWarning)

from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from utils.notification import Notification

class SafelaunchImportRecordAdmin(object):
    list_display = ["user","import_product_name", "import_product_phase", "import_date"]
    list_filter = ["user","import_product_name", "import_product_phase"]
    search_fields = ["user","import_product_name", "import_product_phase"]

class SymptomCategoryFirstAdmin(object):
    list_display = ["category_Name","category_Desc", "cratedate"]
    list_filter = ["category_Name","category_Desc", "cratedate"]
    search_fields = ["category_Name","category_Desc", "cratedate"]

class SymptomCategorySecondAdmin(object):
    list_display = ["category_Name","category_Desc","cratedate"]
    list_filter = ["category_Name","cratedate"]
    search_fields = ["category_Name","category_Desc", "cratedate"]

# define safelaunch issue user interface
class IssueAdmin(object):
    list_display = ["colorStatus","buildstage","platformName","processName","priority","custom_biz_impact","custom_issue_description","impact_scope", "root_cause_category", "input_qty","defect_qty"]
    list_filter = [("platformName__ProductName", MultiSelectFieldListFilter),
                   ("platformName__PartnerName", MultiSelectFieldListFilter),
                   ("processName", MultiSelectFieldListFilter),
                   ("root_cause_category", MultiSelectFieldListFilter),
                   ("long_term_category", MultiSelectFieldListFilter),
                   ("impact_scope", MultiSelectFieldListFilter),
                   ("priority", MultiSelectFieldListFilter),
                   ("buildstage", MultiSelectFieldListFilter),
                   ("status", MultiSelectFieldListFilter),
                   ("cratedate"),
                   ]
    search_fields = ["platformName__ProductName", "issue_desc", "impact_scope","buildstage","owner","status","processName"]
    list_display_links = ["custom_issue_description"]
    list_per_page = 20
    date_hierarchy = 'cratedate'
    # list_editable = ('issue_desc')
    model_icon = 'fa fa-list' #fa fa-cogs

    # configuration for import and export data
    class Issue_import_export_Resource(resources.ModelResource):
        # 外键可视化导出， 而不是导出id
        platformName = fields.Field(
            column_name='platform_name',
            attribute='platformName',
            widget=ForeignKeyWidget(Products, 'ProductName')
        )
        issue_interaction = fields.Field(
            column_name='issue_interaction',
            attribute='issue_interaction',  # issue_interaction 在本模型外键的字段名称
            widget=ForeignKeyWidget(SymptomCategory_First, 'category_Name')  # category_Name 外键里面的字段名
        )
        issue_symptom = fields.Field(
            column_name='issue_symptom',
            attribute='issue_symptom',  # issue_symptom 在本模型外键的字段名称
            widget=ForeignKeyWidget(SymptomCategory_Second, 'category_Name')  # category_Name 外键里面的字段名
        )
        class Meta:
            model = Issue
            exclude = ['id']
            # fields = ('platformName', 'processName', 'issue_desc', 'impact_scope','priority','business_impact','input_qty','defect_qty','issue_analysis') # export feilds
    import_export_args = {'export_resource_class': Issue_import_export_Resource}
    import_excel = True

    # rewrite post method for getting excel safelaunch reprot
    def post(self, request, *args, **kwargs):
        currentuser = self.request.user # get current user

        if 'excel' in request.FILES:
            excel_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(excel_report, data_only=True, keep_vba=False, keep_links=False)

            summary_sheet = workbook['0-Summary Data']
            product_name = summary_sheet.cell(5, 2).value  # get product name from summary sheet
            build_stage = summary_sheet.cell(7, 2).value   # get build stage from summary sheet
            # check if the product is the first upload or repeat to upload(don't support repeated upload)
            record = DataImportRecords.objects.filter(import_product_name=product_name, import_product_phase=build_stage)
            if record:
                # raise a message on webpage
                messages.error(request,
                               "Failed to import report! Error message: [ProductId:{} -- {} -- {}'safelaunch data already uploaded, Repeated upload is not supported]".format(
                                record[0].id, record[0].import_product_name, record[0].import_product_phase))
                # send an alert by email
                contents = "Failed to import report!  Error message: [ProductId:{} -- {} -- {}'safelaunch data already uploaded, Repeated upload is not supported]".format(
                    record[0].id, record[0].import_product_name, record[0].import_product_phase)
                notice = Notification()
                notice.npi_send_by_email(
                                           [str(currentuser)],
                                           currentuser,
                                           contents,
                                           record[0].import_product_name,        # platform name
                                           record[0].import_product_phase,       # build stage
                                           "safelaunch report upload Failed",    # email title
                                           None,
                                           None, )
            else:
                from utils.excel_report_parser import SafelaunchParser
                result = SafelaunchParser().parse(request, workbook, currentuser,)
                # raise a message on webpage
                if len(result) != 1:
                    messages.success(request, "{}-{}'s safe-launch data was successfully uploaded.".format(result[0],result[1]))

                    #   send an alert by email
                    contents = "{}-{} safe-launch data was successfully uploaded.".format(result[0],result[1])
                    Notification().npi_send_by_email(
                                                [str(currentuser)],
                                                currentuser,
                                                contents,
                                                result[0],                                      # platform name
                                                result[1],                                      # build stage
                                                "safelaunch report was successfully uploaded",  # email title
                                                result[3],                                      # 返回data
                                                result[2],)                                     # odm name
                else:
                    messages.error(request,"Failed, error: {}, Please use check tool to check the report first!".format(result[0]))

            workbook.close()
            return HttpResponseRedirect('/scpe/npi/issue/')
        return super(IssueAdmin, self).post(request, *args, **kwargs)

    # configuration pages
    def get_form_layout(self):
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('platformName','buildstage'),
                         Row('status','processName'),
                         Row('priority'),
                         'business_impact',
                         'issue_desc',
                         Row('pre_build_qty', 'pre_build_defcet_qty'),
                         Row('mini_build_qty', 'mini_build_defcet_qty'),
                         Row('mini2_build_qty', 'mini2_build_defcet_qty'),
                         Row('balance_qty', 'balance_defcet_qty'),
                         Row('updatedate', 'owner'),
                         css_class='unsort no_title'
                         ),
                Fieldset(('Basic Information'),
                         Row('issue_interaction', 'issue_symptom'),
                         Row('duplicate', 'errorCode'),
                         'issue_analysis',
                         ),
                Fieldset(('Root Cause & Solutions'),
                         'root_cause',
                         'root_cause_category',
                         'short_term',
                         'short_term_category',
                         'long_term',
                         'long_term_category',
                         ),
                Fieldset((''),
                         Row('fail_rate_stage', 'sku'),
                         ),
                ),
            Side(
                Fieldset(('Key Infomation'),
                         'impact_scope',
                         Row('input_qty', 'defect_qty'),
                         'photo',
                         ),
                )
        )
        return super(IssueAdmin, self).get_form_layout()

# define desktop isksue user interface
class DesktopIssueAdmin(object):
    list_display = ["colorStatus","buildstage","platformName","processName","priority","business_impact","issue_desc","impact_scope", "root_cause_category", "input_qty","defect_qty"]
    list_filter = [("platformName__ProductName", MultiSelectFieldListFilter),
                   ("platformName__PartnerName", MultiSelectFieldListFilter),
                   ("processName", MultiSelectFieldListFilter),
                   ("root_cause_category", MultiSelectFieldListFilter),
                   ("long_term_category", MultiSelectFieldListFilter),
                   ("impact_scope", MultiSelectFieldListFilter),
                   ("priority", MultiSelectFieldListFilter),
                   ("buildstage", MultiSelectFieldListFilter),
                   ("status", MultiSelectFieldListFilter),
                   ("cratedate"),
                   ]
    search_fields = ["platformName__ProductName", "issue_desc", "impact_scope","buildstage","owner","status","processName"]
    list_display_links = ["issue_desc"]
    list_per_page = 20
    date_hierarchy = 'cratedate'
    # list_editable = ('issue_desc')
    model_icon = 'fa fa-list-ol'

    # configuration for import and export data
    class Issue_import_export_Resource(resources.ModelResource):
        # 外键可视化导出， 而不是导入id
        platformName = fields.Field(
            column_name='Platform Name',
            attribute='platformName',
            widget=ForeignKeyWidget(Products, 'ProductName')
        )
        issue_interaction = fields.Field(
            column_name='issue_interaction',
            attribute='issue_interaction',  # issue_interaction 在本模型外键的字段名称
            widget=ForeignKeyWidget(SymptomCategory_First, 'category_Name')  # category_Name 外键里面的字段名
        )
        class Meta:
            model = DesktopIssue
            fields = ('platformName', 'processName', 'issue_desc', 'issue_interaction') # export feilds

    import_export_args = {'export_resource_class': Issue_import_export_Resource}
    import_excel = True

    # rewrite post method for getting excel safelaunch reprot
    def post(self, request, *args, **kwargs):
        currentuser = self.request.user # get current user

        if 'excel' in request.FILES:
            excel_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(excel_report, data_only=True, keep_vba=False, keep_links=False)

            summary_sheet = workbook['0-Summary Data']
            product_name = summary_sheet.cell(5, 2).value  # get product name from summary sheet
            build_stage = summary_sheet.cell(7, 2).value   # get build stage from summary sheet

            # check if the product is the first upload or repeat to upload(don't support repeated upload)
            record = DataImportRecords.objects.filter(import_product_name=product_name, import_product_phase=build_stage)
            if record:
                # raise a message on webpage
                messages.error(request,
                               "Failed to import report! Error message: [ProductId:{} -- {} -- {}'safelaunch data already uploaded, Repeated upload is not supported]".format(
                                record[0].id, record[0].import_product_name, record[0].import_product_phase))

                # send an alert by email
                contents = "Failed to import report!  Error message: [ProductId:{} -- {} -- {}'safelaunch data already uploaded, Repeated upload is not supported]".format(
                    record[0].id, record[0].import_product_name, record[0].import_product_phase)
                notice = Notification()
                notice.npi_send_by_email(
                                           [str(currentuser)],
                                           currentuser,
                                           contents,
                                           record[0].import_product_name,        # platform name
                                           record[0].import_product_phase,       # build stage
                                           "safelaunch report upload Failed",    # email title
                                           None,
                                           None, )

            else:
                from utils.excel_report_parser import DTReportParser
                result = DTReportParser().parse(request, workbook, currentuser,)
                # raise a message on webpage
                if len(result) != 1:
                    messages.success(request, "{}-{}'s safe-launch data was successfully uploaded.".format(result[0],result[1]))

                    """
                    #   send an alert by email
                    contents = "{}-{} safe-launch data was successfully uploaded.".format(result[0],result[1])
                    Notification().npi_send_by_email(
                                                [str(currentuser)],
                                                currentuser,
                                                contents,
                                                result[0],                                      # platform name
                                                result[1],                                      # build stage
                                                "safelaunch report was successfully uploaded",  # email title
                                                result[3],                                      # 返回data
                                                result[2],)                                     # odm name
                    """
                else:
                    messages.error(request,"Failed, error: {}, Please use check tool to check the report first!".format(result[0]))

            workbook.close()
            return HttpResponseRedirect('/scpe/npi/desktopissue/')
        return super(DesktopIssueAdmin, self).post(request, *args, **kwargs)

    # configuration pages
    def get_form_layout(self):
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('platformName','buildstage'),
                         Row('status','processName'),
                         Row('priority'),
                         'business_impact',
                         'issue_desc',
                         # Row('pre_build_qty', 'pre_build_defcet_qty'),
                         Row('mini_build_qty', 'mini_build_defcet_qty'),
                         Row('mini2_build_qty', 'mini2_build_defcet_qty'),
                         Row('balance_qty', 'balance_defcet_qty'),
                         Row('updatedate', 'owner'),
                         css_class='unsort no_title'
                         ),
                Fieldset(('Basic Information'),
                         Row('issue_interaction', 'issue_symptom'),
                         Row('duplicate', 'errorCode'),
                         'issue_analysis',
                         ),
                Fieldset(('Root Cause & Solutions'),
                         'root_cause',
                         'root_cause_category',
                         'short_term',
                         'short_term_category',
                         'long_term',
                         'long_term_category',
                         ),
                Fieldset((''),
                         Row('fail_rate_stage', 'sku'),
                         ),
                ),
            Side(
                Fieldset(('Key Infomation'),
                         'impact_scope',
                         Row('input_qty', 'defect_qty'),
                         'photo',
                         ),
                )
        )
        return super(DesktopIssueAdmin, self).get_form_layout()

# define regional issue user interface
class RegionalCaseAdmin(object):
    list_display = ["issue_status", "issue_desc", "platform_name", "process_name", "build_stage", "input_qty",
                    "defect_qty", "rpe_owner", "gpe_owner",]
    list_filter = [("platform_name__ProductName", MultiSelectFieldListFilter),
                   ("odm_name", MultiSelectFieldListFilter),
                   ("process_name", MultiSelectFieldListFilter),
                   ("root_cause_category", MultiSelectFieldListFilter),
                   ("long_term_category", MultiSelectFieldListFilter),
                   ("impact_scope", MultiSelectFieldListFilter),
                   ("build_stage", MultiSelectFieldListFilter),
                   ("issue_status", MultiSelectFieldListFilter),
                   ("crate_date"),
                   ]
    search_fields = ["platform_name__ProductName", "rpe_owner__username", "gpe_owner", "build_stage", "issue_status",
                     "odm_name","process_name", "impact_scope"]
    list_display_links = ["issue_desc"]
    list_per_page = 20
    date_hierarchy = 'crate_date'
    # list_editable = ('issue_desc')
    model_icon = 'fa fa-bug'

    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         'crate_date',
                         Row('rpe_owner', 'platform_name'),
                         Row('odm_name', 'product_segment',),
                         Row('process_name', 'build_stage'),
                         'issue_desc',

                         css_class='unsort no_title'
                         ),
                Fieldset('',
                         'issue_analysis',
                         'root_cause',
                         'short_term',
                         'long_term',

                         css_class='unsort no_title'
                         ),
            ),
            Side(
                Fieldset('',
                         'gpe_owner',
                         Row('issue_status'),
                         Row('impact_scope'),
                         Row('input_qty', 'defect_qty'),
                         'fail_rate_stage',
                         'issue_photo',

                         css_class='unsort no_title'
                         ),
                Fieldset('',
                         'root_cause_category',
                         'short_term_category',
                         'long_term_category',

                         css_class='unsort no_title'
                         ),
            )
        )
        return super(RegionalCaseAdmin, self).get_form_layout()

# xadmin.site.register(SymptomCategory_First, SymptomCategoryFirstAdmin)
# xadmin.site.register(SymptomCategory_Second, SymptomCategorySecondAdmin)

xadmin.site.register(Issue, IssueAdmin)
# xadmin.site.register(DataImportRecords, SafelaunchImportRecordAdmin)
xadmin.site.register(DesktopIssue, DesktopIssueAdmin)
xadmin.site.register(RegionalCase, RegionalCaseAdmin)